from __future__ import annotations

import re
import sqlite3
import base64
import hashlib
import json
import mimetypes
import subprocess
from datetime import date, datetime
from html import escape
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image as RLImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


APP_DIR = Path(__file__).resolve().parent
ROOT_DIR = APP_DIR.parent
TONG_KET_DIR = ROOT_DIR / "TONG KET"
DB_PATH = APP_DIR / "tong_ket.db"
EXPORT_DIR = APP_DIR / "exports"
UPLOAD_DIR = APP_DIR / "uploads"
BACKUP_DIR = APP_DIR / "backups"
BACKUP_DIR.mkdir(parents=True, exist_ok=True)
ORIGINAL_UPLOAD_DIR = UPLOAD_DIR / "originals"
THUMB_UPLOAD_DIR = UPLOAD_DIR / "thumbs"
MONTH_ORDER = {
    "JANUARY": 1,
    "FEBRUARY": 2,
    "MARCH": 3,
    "APRIL": 4,
    "MAY": 5,
    "JUNE": 6,
    "JULY": 7,
    "AUGUST": 8,
    "SEPTEMBER": 9,
    "OCTOBER": 10,
    "NOVEMBER": 11,
    "DECEMBER": 12,
}
MONTH_NAMES_BY_NUMBER = {value: key for key, value in MONTH_ORDER.items()}


CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS entries (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    source_key TEXT UNIQUE,
    client_name TEXT DEFAULT 'General',
    project_name TEXT NOT NULL,
    owner TEXT,
    period_label TEXT,
    period_year INTEGER,
    description TEXT,
    drawing_qty REAL DEFAULT 0,
    unit_price REAL DEFAULT 0,
    amount REAL DEFAULT 0,
    status TEXT DEFAULT 'Imported',
    notes TEXT,
    image_path TEXT,
    source_file TEXT,
    source_sheet TEXT,
    source_row INTEGER,
    deleted_at TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
)
"""

CREATE_SETTINGS_SQL = """
CREATE TABLE IF NOT EXISTS settings (
    key TEXT PRIMARY KEY,
    value TEXT
)
"""

DEFAULT_PAYMENT_INFO = """Payment methods: Bank transfer
Bank name: Asia Commercial Joint Stock Bank (ACB - A Chau Bank)
Account number: 196653719
Full name: LUONG NHAT TU
Swift code: ASCBVNVX"""




def auto_backup_db() -> "Path | None":
    """Create a daily backup of the DB. Returns backup path if new backup was made."""
    today = datetime.now().strftime("%Y-%m-%d")
    backup_path = BACKUP_DIR / f"tong_ket_{today}.db"
    if backup_path.exists():
        return None
    try:
        import shutil
        shutil.copy2(DB_PATH, backup_path)
        backups = sorted(BACKUP_DIR.glob("tong_ket_*.db"))
        for old in backups[:-30]:
            old.unlink(missing_ok=True)
        return backup_path
    except Exception:
        return None

def connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.execute(CREATE_TABLE_SQL)
    conn.execute(CREATE_SETTINGS_SQL)
    ensure_schema(conn)
    return conn


def ensure_schema(conn: sqlite3.Connection) -> None:
    columns = {row[1] for row in conn.execute("PRAGMA table_info(entries)").fetchall()}
    if "client_name" not in columns:
        conn.execute("ALTER TABLE entries ADD COLUMN client_name TEXT DEFAULT 'General'")
    if "deleted_at" not in columns:
        conn.execute("ALTER TABLE entries ADD COLUMN deleted_at TEXT")
    if "image_path" not in columns:
        conn.execute("ALTER TABLE entries ADD COLUMN image_path TEXT")
    if conn.execute("SELECT value FROM settings WHERE key = 'payment_info'").fetchone() is None:
        conn.execute("INSERT INTO settings (key, value) VALUES (?, ?)", ("payment_info", DEFAULT_PAYMENT_INFO))
    if conn.execute("SELECT value FROM settings WHERE key = 'import_folder'").fetchone() is None:
        conn.execute("INSERT INTO settings (key, value) VALUES (?, ?)", ("import_folder", str(TONG_KET_DIR)))


def get_setting(key: str, default: str = "") -> str:
    with connect() as conn:
        row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    return str(row[0]) if row and row[0] is not None else default


def set_setting(key: str, value: str) -> None:
    with connect() as conn:
        conn.execute(
            "INSERT INTO settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value",
            (key, value),
        )


def unique_clean(values: list[object]) -> list[str]:
    return sorted({normalize_text(value) for value in values if normalize_text(value)}, key=str.lower)


def default_company_owner_config(data: pd.DataFrame | None = None) -> dict[str, object]:
    companies = ["ERNEST", "ETHAN", "JOSELYN", "KING'S CARPENTRY"]
    owners_by_company: dict[str, list[str]] = {company: [company] for company in companies}
    if data is not None and not data.empty:
        companies = unique_clean([uppercase_label(value) for value in data["client_name"].fillna("GENERAL").tolist()] + companies)
        owners_by_company = {company: owners_by_company.get(company, [company]) for company in companies}
        for company, rows in data.groupby("client_name", dropna=False):
            company_name = uppercase_label(company) or "GENERAL"
            owners = unique_clean([uppercase_label(value) for value in rows["owner"].dropna().tolist()])
            if owners:
                owners_by_company[company_name] = owners
    return {"companies": companies, "owners_by_company": owners_by_company}


def get_company_owner_config(data: pd.DataFrame | None = None) -> dict[str, object]:
    raw = get_setting("company_owner_config", "")
    fallback = default_company_owner_config(data)
    if not raw.strip():
        return fallback
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        return fallback
    companies = unique_clean([uppercase_label(value) for value in list(parsed.get("companies") or []) + list(fallback.get("companies") or [])])
    owners_by_company = dict(fallback.get("owners_by_company") or {})
    for company, owners in (parsed.get("owners_by_company") or {}).items():
        company_name = uppercase_label(company)
        if company_name:
            owners_by_company[company_name] = unique_clean([uppercase_label(value) for value in list(owners or [])])
    return {"companies": companies, "owners_by_company": owners_by_company}


def save_company_owner_config(companies: list[str], owners_by_company: dict[str, list[str]]) -> None:
    clean_companies = unique_clean([uppercase_label(value) for value in companies])
    clean_owners = {company: unique_clean([uppercase_label(value) for value in owners_by_company.get(company, [])]) for company in clean_companies}
    set_setting(
        "company_owner_config",
        json.dumps({"companies": clean_companies, "owners_by_company": clean_owners}, ensure_ascii=False, indent=2),
    )


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def uppercase_project_name(value: object) -> str:
    return normalize_text(value).upper()


def uppercase_label(value: object) -> str:
    return normalize_text(value).upper()


def number_or_zero(value: object) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def clean_project_text(text: str) -> str:
    return "\n".join(line.strip() for line in normalize_text(text).splitlines() if line.strip())


def split_project(project_text: str) -> tuple[str, str, str]:
    lines = clean_project_text(project_text).splitlines()
    if not lines:
        return "", "", ""

    title = lines[0].strip()
    description_parts = lines[1:]
    inline_match = re.match(r"^(.*?)\s*\((.*)\)\s*$", title)
    if inline_match:
        title = inline_match.group(1).strip()
        description_parts.insert(0, inline_match.group(2).strip())

    owner = ""
    if " - " in title:
        candidate, owner_candidate = title.rsplit(" - ", 1)
        if owner_candidate.strip():
            title = candidate.strip()
            owner = uppercase_label(owner_candidate)

    description = "\n".join(description_parts).strip()
    return uppercase_project_name(title), owner, description


def parse_year(text: str) -> int | None:
    match = re.search(r"(20\d{2})", text or "")
    return int(match.group(1)) if match else None


def normalize_period_label(value: object) -> str:
    if isinstance(value, (datetime, date)):
        return f"{MONTH_NAMES_BY_NUMBER[value.month]} - {value.year}"
    return normalize_text(value)


def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return cleaned.strip("_") or "project_summary"


def detect_client_name(path: Path, root: Path) -> str:
    relative_parts = path.relative_to(root).parts
    top_folder = normalize_text(relative_parts[0]).upper() if len(relative_parts) > 1 else ""
    if top_folder and top_folder not in {"CARPENTRY", "OLD", "HOA DON"}:
        return top_folder

    stem = normalize_text(path.stem)
    dash_match = re.match(r"^([A-Za-z][A-Za-z0-9 &._']*)\s*-\s+", stem)
    if dash_match:
        return dash_match.group(1).strip().upper()

    summary_match = re.search(r"\b(ERNEST|RYAN|KELVIN)\b", stem, re.IGNORECASE)
    if summary_match:
        return summary_match.group(1).upper()

    for part in path.relative_to(root).parts[:-1]:
        if re.search(r"\b(ERNEST|RYAN|KELVIN)\b", part, re.IGNORECASE):
            return re.search(r"\b(ERNEST|RYAN|KELVIN)\b", part, re.IGNORECASE).group(1).upper()
    return "GENERAL"


def detect_period_from_path(path: Path, root: Path) -> str:
    candidates = [path.stem, *reversed(path.relative_to(root).parts[:-1])]
    for candidate in candidates:
        text = normalize_text(candidate).upper()
        year = parse_year(text)
        if not year:
            continue
        months = [month_name for month_name in MONTH_ORDER if month_name in text]
        if months:
            return f"{' + '.join(months)} - {year}"
        numeric_match = re.search(r"(?<!\d)(1[0-2]|0?[1-9])\s*[-.]\s*(20\d{2})", text)
        if numeric_match:
            month_number = int(numeric_match.group(1))
            return f"{MONTH_NAMES_BY_NUMBER[month_number]} - {numeric_match.group(2)}"
    return ""


def extract_sheet_images(path: Path, sheet, client_name: str) -> dict[int, list[str]]:
    image_map: dict[int, list[str]] = {}
    images = list(getattr(sheet, "_images", []) or [])
    if not images:
        return image_map

    image_root = UPLOAD_DIR / "excel_imports" / safe_filename(client_name) / safe_filename(path.stem)
    image_root.mkdir(parents=True, exist_ok=True)
    sheet_name = safe_filename(sheet.title)
    for index, image in enumerate(images, start=1):
        marker = getattr(getattr(image, "anchor", None), "_from", None)
        if marker is None:
            continue
        row_number = int(marker.row) + 1
        output_path = image_root / f"{sheet_name}_r{row_number:04d}_{index:03d}.jpg"
        if not output_path.exists():
            try:
                with Image.open(BytesIO(image._data())) as pil_image:
                    pil_image = pil_image.convert("RGB")
                    pil_image.thumbnail((640, 640))
                    pil_image.save(output_path, format="JPEG", quality=82, optimize=True)
            except Exception:
                continue
        image_map.setdefault(row_number, []).append(str(output_path))
    return image_map


def is_header_row(row: tuple[object, ...]) -> bool:
    values = [normalize_text(v).lower() for v in row[:5]]
    return "projects" in values and any(("drawing" in v or "view" in v or "quantity" in v) for v in values)


def is_period_row(row: tuple[object, ...]) -> bool:
    first = normalize_text(row[0] if row else "")
    second = normalize_text(row[1] if len(row) > 1 else "")
    if not first or second:
        return False
    upper = first.upper()
    return bool(re.search(r"20\d{2}", upper)) and "TOTAL" not in upper


def excel_files_for_import(root: Path) -> list[Path]:
    return [
        path
        for path in sorted(root.rglob("*.xlsx"), key=lambda item: str(item).lower())
        if not path.name.startswith("~$") and path.name != "TONG_KET_MASTER_INDEX.xlsx"
    ]


def build_source_key(path: Path, root: Path, sheet_title: str, row_index: int) -> str:
    try:
        path_key = str(path.relative_to(root))
    except ValueError:
        path_key = str(path)
    return f"{path_key}|{sheet_title}|{row_index}"


def migrate_absolute_image_paths() -> int:
    """Chuyển image_path từ tuyệt đối Windows -> relative so với APP_DIR.
    Chạy 1 lần để sửa các bản ghi đã import trước đó."""
    migrated = 0
    with connect() as conn:
        rows = conn.execute(
            "SELECT id, image_path FROM entries "
            "WHERE image_path LIKE '%\\%' OR image_path LIKE 'D:%' "
            "OR image_path LIKE 'C:%' OR image_path LIKE 'E:%'"
        ).fetchall()
        for entry_id, image_path in rows:
            if not image_path:
                continue
            new_parts: list[str] = []
            changed = False
            for part in image_path.split("|"):
                part = part.strip()
                if not part:
                    continue
                normalized = part.replace("\\", "/")
                p = Path(normalized)
                parts_lower = [x.lower() for x in p.parts]
                if "uploads" in parts_lower:
                    idx = parts_lower.index("uploads")
                    rel = "/".join(p.parts[idx + 1:])
                    new_parts.append(f"uploads/{rel}")
                    changed = True
                else:
                    new_parts.append(part)
            if changed:
                new_value = "|".join(new_parts)
                conn.execute("UPDATE entries SET image_path = ? WHERE id = ?", (new_value, entry_id))
                migrated += 1
    if migrated:
        clear_entries_cache()
    return migrated

def discover_excel_rows(
    root: Path,
    errors: list[str] | None = None,
    files: list[Path] | None = None,
    progress_callback: object | None = None,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    excel_files = files if files is not None else excel_files_for_import(root)
    total_files = len(excel_files)
    for file_index, path in enumerate(excel_files, start=1):
        if progress_callback:
            progress_callback(file_index, total_files, path)
        try:
            workbook = load_workbook(path, read_only=False, data_only=True)
        except Exception as exc:
            if errors is not None:
                errors.append(f"{path}: {type(exc).__name__}: {exc}")
            continue

        client_name = detect_client_name(path, root)
        fallback_period = detect_period_from_path(path, root)
        for sheet in workbook.worksheets:
            image_map = extract_sheet_images(path, sheet, client_name)
            header_seen = False
            current_period = fallback_period
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if is_period_row(row):
                    current_period = normalize_period_label(row[0])
                    continue
                if is_header_row(row):
                    header_seen = True
                    continue
                if not header_seen:
                    continue

                first = normalize_text(row[0] if row else "")
                project_text = clean_project_text(row[1] if len(row) > 1 else "")
                if not project_text:
                    continue
                lowered_project = project_text.lower()
                if first.upper().startswith("TOTAL") or lowered_project.startswith("payment") or lowered_project in {
                    "my info banking",
                    "bank name",
                    "account number",
                    "full name (payee name)",
                    "swiftcode",
                }:
                    continue
                drawing_qty = number_or_zero(row[2] if len(row) > 2 else 0)
                unit_price = number_or_zero(row[3] if len(row) > 3 else 0)
                amount = number_or_zero(row[4] if len(row) > 4 else 0)
                if drawing_qty == 0 and amount == 0:
                    continue
                has_row_number = bool(re.fullmatch(r"\d+(?:\.0+)?", first))
                has_project_values = bool(project_text and (drawing_qty != 0 or amount != 0))
                if not has_row_number and not has_project_values:
                    continue

                project_name, owner, description = split_project(project_text)
                if not project_name:
                    continue

                source_key = build_source_key(path, root, sheet.title, row_index)
                image_paths = image_map.get(row_index, [])
                rows.append(
                    {
                        "source_key": source_key,
                        "client_name": client_name,
                        "project_name": project_name,
                        "owner": owner,
                        "period_label": current_period,
                        "period_year": parse_year(current_period),
                        "description": description,
                        "drawing_qty": drawing_qty,
                        "unit_price": unit_price,
                        "amount": amount,
                        "status": "Imported",
                        "notes": "",
                        "image_path": "|".join(image_paths),
                        "source_file": str(path),
                        "source_sheet": sheet.title,
                        "source_row": row_index,
                    }
                )
        workbook.close()
    return rows


def import_excel_data(root: Path | None = None, progress_callback: object | None = None) -> int:
    now = datetime.now().isoformat(timespec="seconds")
    import_root = root or TONG_KET_DIR
    errors: list[str] = []
    excel_files = excel_files_for_import(import_root)
    files_found = len(excel_files)
    rows = discover_excel_rows(import_root, errors, files=excel_files, progress_callback=progress_callback)
    with connect() as conn:
        migrated_keys = migrate_absolute_source_keys(conn, import_root)
        conn.execute("CREATE TEMP TABLE IF NOT EXISTS current_import_keys (source_key TEXT PRIMARY KEY)")
        conn.execute("DELETE FROM current_import_keys")
        conn.executemany(
            "INSERT OR IGNORE INTO current_import_keys (source_key) VALUES (?)",
            [(row["source_key"],) for row in rows],
        )
        for row in rows:
            conn.execute(
                """
                INSERT INTO entries (
                    source_key, client_name, project_name, owner, period_label, period_year,
                    description, drawing_qty, unit_price, amount, status, notes, image_path,
                    source_file, source_sheet, source_row, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(source_key) DO UPDATE SET
                    client_name=excluded.client_name,
                    project_name=excluded.project_name,
                    owner=excluded.owner,
                    period_label=excluded.period_label,
                    period_year=excluded.period_year,
                    description=excluded.description,
                    drawing_qty=excluded.drawing_qty,
                    unit_price=excluded.unit_price,
                    amount=excluded.amount,
                    image_path=CASE
                        WHEN entries.status = 'Imported' THEN excluded.image_path
                        WHEN COALESCE(entries.image_path, '') = '' THEN excluded.image_path
                        ELSE entries.image_path
                    END,
                    source_file=excluded.source_file,
                    source_sheet=excluded.source_sheet,
                    source_row=excluded.source_row,
                    updated_at=excluded.updated_at
                WHERE entries.status = 'Imported'
                """,
                (
                    row["source_key"],
                    row["client_name"],
                    row["project_name"],
                    row["owner"],
                    row["period_label"],
                    row["period_year"],
                    row["description"],
                    row["drawing_qty"],
                    row["unit_price"],
                    row["amount"],
                    row["status"],
                    row["notes"],
                    row["image_path"],
                    row["source_file"],
                    row["source_sheet"],
                    row["source_row"],
                    now,
                    now,
                ),
            )
        cursor = conn.execute(
            """
            UPDATE entries
            SET deleted_at = ?, updated_at = ?
            WHERE status = 'Imported'
              AND deleted_at IS NULL
              AND source_key NOT LIKE 'manual|%'
              AND source_key NOT IN (SELECT source_key FROM current_import_keys)
            """,
            (now, now),
        )
        rows_deleted = int(conn.execute("SELECT changes()").fetchone()[0])
        summary = {
            "imported_at": now,
            "import_folder": str(import_root),
            "files_found": files_found,
            "rows_discovered": len(rows),
            "rows_upserted": len(rows),
            "rows_deleted": rows_deleted,
            "source_keys_migrated": migrated_keys,
            "errors": errors,
        }
        conn.execute(
            "INSERT INTO settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value",
            ("last_import_summary", json.dumps(summary, ensure_ascii=False, indent=2)),
        )
    clear_entries_cache()
    return len(rows)


@st.cache_data(show_spinner=False)
def load_entries() -> pd.DataFrame:
    with connect() as conn:
        data = pd.read_sql_query(
            "SELECT * FROM entries WHERE deleted_at IS NULL ORDER BY client_name, period_year DESC, id DESC",
            conn,
        )
    return data


def clear_entries_cache() -> None:
    try:
        load_entries.clear()
    except Exception:
        pass


def get_entry(entry_id: int) -> dict[str, object] | None:
    with connect() as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM entries WHERE id = ? AND deleted_at IS NULL", (entry_id,)).fetchone()
    return dict(row) if row else None


def update_entry(entry_id: int, values: dict[str, object]) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as conn:
        conn.execute(
            """
            UPDATE entries
            SET client_name = ?,
                project_name = ?,
                owner = ?,
                period_label = ?,
                period_year = ?,
                description = ?,
                drawing_qty = ?,
                unit_price = ?,
                amount = ?,
                notes = ?,
                image_path = ?,
                status = CASE WHEN status = 'Manual' THEN 'Manual' ELSE 'Edited' END,
                updated_at = ?
            WHERE id = ? AND deleted_at IS NULL
            """,
            (
                uppercase_label(values.get("client_name", "GENERAL")) or "GENERAL",
                uppercase_project_name(values["project_name"]),
                uppercase_label(values.get("owner", "")),
                values.get("period_label", ""),
                parse_year(str(values.get("period_label", ""))),
                values.get("description", ""),
                values.get("drawing_qty", 0),
                values.get("unit_price", 0),
                values.get("amount", 0),
                values.get("notes", ""),
                values.get("image_path", ""),
                now,
                entry_id,
            ),
        )
    clear_entries_cache()


def soft_delete_entry(entry_id: int) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as conn:
        conn.execute(
            "UPDATE entries SET deleted_at = ?, updated_at = ? WHERE id = ? AND deleted_at IS NULL",
            (now, now, entry_id),
        )
    clear_entries_cache()


def soft_delete_entries(entry_ids: list[int]) -> int:
    ids = [int(entry_id) for entry_id in entry_ids if entry_id]
    if not ids:
        return 0
    now = datetime.now().isoformat(timespec="seconds")
    placeholders = ",".join("?" for _ in ids)
    with connect() as conn:
        cursor = conn.execute(
            f"UPDATE entries SET deleted_at = ?, updated_at = ? WHERE deleted_at IS NULL AND id IN ({placeholders})",
            (now, now, *ids),
        )
        deleted_count = int(conn.execute("SELECT changes()").fetchone()[0])
    clear_entries_cache()
    return deleted_count




def restore_entry(entry_id: int) -> None:
    """Restore a soft-deleted entry."""
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as conn:
        conn.execute(
            "UPDATE entries SET deleted_at = NULL, updated_at = ? WHERE id = ?",
            (now, entry_id),
        )
    clear_entries_cache()


def load_deleted_entries() -> "pd.DataFrame":
    """Load soft-deleted entries for trash bin view."""
    with connect() as conn:
        data = pd.read_sql_query(
            "SELECT * FROM entries WHERE deleted_at IS NOT NULL ORDER BY deleted_at DESC",
            conn,
        )
    return data


def permanently_delete_entry(entry_id: int) -> None:
    """Hard delete a single entry."""
    with connect() as conn:
        conn.execute("DELETE FROM entries WHERE id = ?", (entry_id,))
    clear_entries_cache()

def insert_manual_entry(values: dict[str, object]) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    project_name = uppercase_project_name(values["project_name"])
    source_key = f"manual|{now}|{project_name}|{id(values)}"
    with connect() as conn:
        conn.execute(
            """
            INSERT INTO entries (
                source_key, client_name, project_name, owner, period_label, period_year,
                description, drawing_qty, unit_price, amount, status, notes, image_path,
                source_file, source_sheet, source_row, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                source_key,
                uppercase_label(values.get("client_name", "GENERAL")) or "GENERAL",
                project_name,
                uppercase_label(values.get("owner", "")),
                values.get("period_label", ""),
                parse_year(str(values.get("period_label", ""))),
                values.get("description", ""),
                values.get("drawing_qty", 0),
                values.get("unit_price", 0),
                values.get("amount", 0),
                "Manual",
                values.get("notes", ""),
                values.get("image_path", ""),
                "",
                "",
                None,
                now,
                now,
            ),
        )
    clear_entries_cache()


def insert_manual_entries(entries: list[dict[str, object]]) -> int:
    inserted = 0
    for entry in entries:
        project_name = uppercase_project_name(entry.get("project_name"))
        if not project_name:
            continue
        insert_manual_entry({**entry, "project_name": project_name})
        inserted += 1
    return inserted


def filter_entries(data: pd.DataFrame, keyword: str, client: str, owner: str, year: str, period: str) -> pd.DataFrame:
    if data.empty:
        return data
    filtered = data.copy()
    if keyword:
        keyword_lower = keyword.lower()
        text_cols = ["client_name", "project_name", "owner", "period_label", "description", "notes"]
        mask = filtered[text_cols].fillna("").agg(" ".join, axis=1).str.lower().str.contains(keyword_lower, regex=False)
        filtered = filtered[mask]
    if client != "All":
        filtered = filtered[filtered["client_name"].fillna("GENERAL").map(uppercase_label) == uppercase_label(client)]
    if owner != "All":
        filtered = filtered[filtered["owner"].fillna("").map(uppercase_label) == uppercase_label(owner)]
    if year != "All":
        filtered = filtered[filtered["period_year"].fillna(0).astype(int).astype(str) == year]
    if period != "All":
        filtered = filtered[filtered["period_label"].fillna("") == period]
    return filtered


def search_autocomplete_options(data: pd.DataFrame, query: str, limit: int = 10) -> list[tuple[str, str]]:
    query_text = normalize_text(query).lower()
    if data.empty or not query_text:
        return []
    buckets = [
        ("Dự án", "project_name"),
        ("Công ty", "client_name"),
        ("Người phụ trách", "owner"),
        ("Kỳ", "period_label"),
    ]
    matches: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for label, column in buckets:
        source_values = [uppercase_project_name(value) for value in data[column].dropna().tolist()] if column == "project_name" else [uppercase_label(value) for value in data[column].dropna().tolist()]
        for value in unique_clean(source_values):
            value_text = str(value).strip()
            if not value_text or query_text not in value_text.lower():
                continue
            key = (label, value_text.lower())
            if key in seen:
                continue
            seen.add(key)
            matches.append((f"{label} · {value_text}", value_text))
            if len(matches) >= limit:
                return matches
    return matches


def period_sort_key(period_label: str) -> tuple[int, str]:
    year = parse_year(period_label) or 0
    upper = normalize_text(period_label).upper()
    months = [month_number for month_name, month_number in MONTH_ORDER.items() if month_name in upper]
    month = max(months) if months else 0
    return (year, month, upper)


def year_options(data: pd.DataFrame) -> list[str]:
    if data.empty:
        return []
    years = sorted([str(int(x)) for x in data["period_year"].dropna().unique().tolist()], reverse=True)
    return years


def periods_for_year(data: pd.DataFrame, year: str) -> list[str]:
    if data.empty:
        return []
    scoped = data
    if year != "All":
        scoped = data[data["period_year"].fillna(0).astype(int).astype(str) == year]
    return sorted(
        [x for x in scoped["period_label"].dropna().unique().tolist() if x],
        key=period_sort_key,
        reverse=True,
    )


def project_history_options(data: pd.DataFrame, client_name: str) -> list[str]:
    if data.empty:
        return []
    scoped = data[data["client_name"].fillna("GENERAL").map(uppercase_label) == (uppercase_label(client_name) or "GENERAL")]
    return unique_clean([uppercase_project_name(value) for value in scoped["project_name"].dropna().tolist()])


def project_history_summary(data: pd.DataFrame, client_name: str, project_name: str) -> dict[str, object] | None:
    if data.empty or not project_name:
        return None
    scoped = data[
        (data["client_name"].fillna("GENERAL").map(uppercase_label) == (uppercase_label(client_name) or "GENERAL"))
        & (data["project_name"].fillna("").map(uppercase_project_name) == uppercase_project_name(project_name))
    ].copy()
    if scoped.empty:
        return None
    scoped["_period_sort"] = scoped["period_label"].fillna("").map(period_sort_key)
    first = scoped.sort_values(["period_year", "_period_sort", "created_at", "id"], ascending=[True, True, True, True]).iloc[0]
    latest = scoped.sort_values(["updated_at", "id"], ascending=[False, False]).iloc[0]
    scoped_sorted = scoped.sort_values(["period_year", "_period_sort", "created_at", "id"], ascending=[True, True, True, True])
    return {
        "first_period": str(first.get("period_label") or ""),
        "latest_owner": uppercase_label(latest.get("owner")),
        "latest_unit_price": number_or_zero(latest.get("unit_price")),
        "rows": scoped_sorted.drop(columns=["_period_sort"], errors="ignore").to_dict("records"),
    }


def client_options(data: pd.DataFrame) -> list[str]:
    if data.empty or "client_name" not in data.columns:
        return []
    return sorted([x for x in unique_clean([uppercase_label(value) for value in data["client_name"].fillna("GENERAL").tolist()]) if x])


def excel_like_period_table(data: pd.DataFrame) -> pd.DataFrame:
    if data.empty:
        return pd.DataFrame(columns=["No", "Project", "Image", "Drawings Quantity", "Unit Price (SGD)", "Amount (SGD)"])

    period_data = data.sort_values(["source_file", "source_row", "id"], na_position="last").reset_index(drop=True)
    rows = []
    for index, row in period_data.iterrows():
        title = uppercase_project_name(row.get("project_name"))
        owner = uppercase_label(row.get("owner"))
        description = str(row.get("description") or "").strip()
        project_text = f"{title} - {owner}" if owner else title
        if description:
            project_text = f"{project_text}\n{description}"
        rows.append(
            {
                "No": index + 1,
                "Project": project_text,
                "Image": str(row.get("image_path") or "").strip(),
                "Drawings Quantity": number_or_zero(row.get("drawing_qty")),
                "Unit Price (SGD)": number_or_zero(row.get("unit_price")),
                "Amount (SGD)": number_or_zero(row.get("amount")),
            }
        )

    total = {
        "No": "",
        "Project": f"TOTAL IN {str(period_data.iloc[0].get('period_label') or '').upper()}",
        "Image": "",
        "Drawings Quantity": period_data["drawing_qty"].fillna(0).sum(),
        "Unit Price (SGD)": "",
        "Amount (SGD)": period_data["amount"].fillna(0).sum(),
    }
    rows.append(total)
    return pd.DataFrame(rows)


def format_multiline_text(value: object) -> str:
    text = normalize_text(value)
    text = re.sub(r"\s+\*", "\n*", text)
    text = re.sub(r"\s+-\s+", " - ", text)
    return "\n".join(line.strip() for line in text.splitlines() if line.strip())


def resolve_image_path(image_path: str) -> Path:
    path = Path(image_path)
    if not path.is_absolute():
        path = APP_DIR / path
    return path


@st.cache_data(show_spinner=False)
def _image_data_uri_cached(path_str: str, mtime: float) -> str:
    path = Path(path_str)
    if not path.exists() or not path.is_file():
        return ""
    mime_type = mimetypes.guess_type(str(path))[0] or "image/jpeg"
    data = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime_type};base64,{data}"


def image_data_uri(image_path: object) -> str:
    path_text = normalize_text(image_path)
    if not path_text:
        return ""
    path = resolve_image_path(path_text)
    if not path.exists() or not path.is_file():
        return ""
    return _image_data_uri_cached(str(path), path.stat().st_mtime)


def original_image_path_for_thumb(image_path: object) -> Path:
    path_text = normalize_text(image_path)
    path = resolve_image_path(path_text)
    try:
        relative = path.relative_to(UPLOAD_DIR)
    except ValueError:
        return path
    if not relative.parts or relative.parts[0] != "thumbs":
        return path

    original_dir = UPLOAD_DIR / "originals"
    matches = sorted(original_dir.glob(f"{path.stem}.*"))
    return matches[0] if matches else path


def original_image_data_uri(image_path: object) -> str:
    original_path = original_image_path_for_thumb(image_path)
    return image_data_uri(str(original_path))


def image_paths_from_value(value: object) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    return [part.strip() for part in text.split("|") if part.strip()]


def render_image_thumbnails(value: object) -> str:
    image_tags = []
    for path_text in image_paths_from_value(value):
        image_uri = image_data_uri(path_text)
        original_uri = original_image_data_uri(path_text)
        if image_uri:
            image_tags.append(
                f'<img class="ref-image image-open" src="{image_uri}" data-full="{escape(original_uri or image_uri, quote=True)}" />'
            )
    return "".join(image_tags)


def compact_started_text(value: str) -> str:
    text = normalize_text(value)
    match = re.search(r"(?:project\s+)?started\s+in\s+(.+)$", text, re.IGNORECASE)
    if not match:
        return text
    period_text = match.group(1).strip().strip("\"'()")
    return f"started in {period_text}"


def split_period_project_text(title: object, owner: object, description: object) -> tuple[str, str, list[str], str]:
    project_title = uppercase_project_name(title)
    owner_text = uppercase_label(owner)
    started_text = ""
    description_lines: list[str] = []

    for line in format_multiline_text(description).splitlines():
        clean_line = re.sub(r"^[*\-]\s*", "", line).strip()
        if not clean_line:
            continue
        if re.search(r"(?:project\s+)?started\s+in\s+", clean_line, re.IGNORECASE):
            started_text = compact_started_text(clean_line)
            continue
        description_lines.append(clean_line)

    owner_line = owner_text if owner_text else ""
    if owner_line and started_text:
        owner_line = f"{owner_line} ({started_text})"
    elif started_text:
        owner_line = f"({started_text})"
    return project_title, owner_line, description_lines, started_text


def render_period_image_grid(value: object, max_images: int = 8) -> str:
    paths = image_paths_from_value(value)
    if not paths:
        return ""
    visible_paths = paths[:max_images]
    extra_count = max(0, len(paths) - max_images)
    cells = []
    for path_text in visible_paths:
        image_uri = image_data_uri(path_text)
        original_uri = original_image_data_uri(path_text)
        if image_uri:
            cells.append(
                f"""
                <button type="button" class="image-cell image-open" data-full="{escape(original_uri or image_uri, quote=True)}">
                    <img src="{image_uri}" />
                </button>
                """
            )
    if extra_count:
        cells.append(f'<div class="image-cell image-more">+{extra_count}</div>')
    return f'<div class="image-grid">{"".join(cells)}</div>' if cells else ""


def pdf_image_grid(value: object, max_images: int = 12) -> Table | str:
    image_cells: list[object] = []
    for path_text in image_paths_from_value(value)[:max_images]:
        path = resolve_image_path(path_text)
        if path.exists():
            image = RLImage(str(path))
            image._restrictSize(16 * mm, 16 * mm)
            image_cells.append(image)
    if not image_cells:
        return ""
    rows = []
    for index in range(0, len(image_cells), 3):
        row = image_cells[index : index + 3]
        rows.append(row + [""] * (3 - len(row)))
    grid = Table(rows, colWidths=[16 * mm, 16 * mm, 16 * mm], rowHeights=[16 * mm] * len(rows))
    grid.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 1),
                ("RIGHTPADDING", (0, 0), (-1, -1), 1),
                ("TOPPADDING", (0, 0), (-1, -1), 1),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ]
        )
    )
    return grid


def calc_pdf_row_height(image_value: object, text_lines: list[str], include_period: bool = False) -> float:
    image_count = min(len(image_paths_from_value(image_value)), 12)
    image_rows = (image_count + 2) // 3 if image_count else 0
    height_from_images = (image_rows * 18 * mm) + (4 * mm) if image_rows else 0

    wrapped_text_lines = 0
    for line in text_lines:
        text = normalize_text(line)
        wrapped_text_lines += max(1, (len(text) // (34 if include_period else 46)) + 1) if text else 1
    height_from_text = max(1, wrapped_text_lines) * 4.2 * mm + (5 * mm)
    return max(16 * mm, height_from_images, height_from_text)


def paginate_dataframe(data: pd.DataFrame, key: str, page_size: int = 50) -> pd.DataFrame:
    if data.empty:
        return data
    total_rows = len(data)
    if total_rows <= page_size:
        return data
    
    total_pages = (total_rows - 1) // page_size + 1
    page_key = f"{key}_page_num"
    if page_key not in st.session_state:
        st.session_state[page_key] = 1
    
    # Reset page if it's out of bounds (e.g. after filtering)
    if st.session_state[page_key] > total_pages:
        st.session_state[page_key] = 1

    cols = st.columns([2, 3, 2])
    with cols[1]:
        sub_cols = st.columns([1, 2, 1])
        if sub_cols[0].button("←", key=f"{key}_prev", disabled=st.session_state[page_key] <= 1, width="stretch"):
            st.session_state[page_key] -= 1
            st.rerun()
        sub_cols[1].markdown(f"<div style='text-align:center; padding-top:5px; font-size:14px; color:var(--muted);'>Trang <b>{st.session_state[page_key]}</b> / {total_pages}</div>", unsafe_allow_html=True)
        if sub_cols[2].button("→", key=f"{key}_next", disabled=st.session_state[page_key] >= total_pages, width="stretch"):
            st.session_state[page_key] += 1
            st.rerun()
            
    start_idx = (st.session_state[page_key] - 1) * page_size
    end_idx = min(start_idx + page_size, total_rows)
    return data.iloc[start_idx:end_idx]


def empty_state(message: str, icon: str = "📦", subtitle: str = "") -> None:
    st.markdown(
        f"""
        <div style="display: flex; flex-direction: column; align-items: center; justify-content: center; padding: 60px 20px; background: var(--bg); border: 2px dashed var(--border); border-radius: 16px; color: var(--muted); text-align: center; margin: 10px 0;">
            <div style="font-size: 48px; margin-bottom: 16px; filter: grayscale(0.5);">{icon}</div>
            <div style="font-size: 18px; font-weight: 600; color: var(--text);">{escape(message)}</div>
            {f'<div style="font-size: 14px; margin-top: 8px;">{escape(subtitle)}</div>' if subtitle else ""}
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_period_table(data: pd.DataFrame) -> None:
    import uuid
    if data.empty:
        empty_state("Không có dữ liệu trong kỳ này", icon="🔍", subtitle="Hãy thử chọn kỳ khác hoặc import thêm dữ liệu.")
        return

    wrap_id = "tw_" + uuid.uuid4().hex[:8]

    # Pagination
    display_data = paginate_dataframe(data, "period_table_page")
    period_data = display_data.sort_values(["source_file", "source_row", "id"], na_position="last").reset_index(drop=True)

    # Build rows
    rows_parts = []
    estimated_body_height = 0
    for index, row in period_data.iterrows():
        project_title, owner_line, description_lines, _started_text = split_period_project_text(
            row.get("project_name"),
            row.get("owner"),
            row.get("description"),
        )
        desc_html = "".join(
            '<div class="pdl"><span class="dot"></span>' + escape(line) + '</div>'
            for line in description_lines
        )
        owner_html = ('<span class="poi"> &middot; ' + escape(owner_line) + '</span>') if owner_line else ""
        image_html = render_period_image_grid(row.get("image_path"))
        qty = f"{number_or_zero(row.get('drawing_qty')):g}"
        unit = f"SGD {number_or_zero(row.get('unit_price')):,.0f}"
        amount = f"SGD {number_or_zero(row.get('amount')):,.0f}"
        text_lines = [project_title + (f" · {owner_line}" if owner_line else "")] + description_lines
        wrapped_text_lines = sum(max(1, (len(line) + 54) // 55) for line in text_lines if line)
        image_count = min(len(image_paths_from_value(row.get("image_path"))), 8)
        image_rows = (image_count + 2) // 3 if image_count else 0
        estimated_body_height += max(72, 34 + wrapped_text_lines * 21, 22 + image_rows * 90)
        rows_parts.append(
            '<tr>'
            + '<td class="cn">' + str(index + 1) + '</td>'
            + '<td class="cp"><div class="pn">' + escape(project_title) + owner_html + '</div><div class="pd">' + desc_html + '</div></td>'
            + '<td class="cq">' + qty + '</td>'
            + '<td class="cu">' + unit + '</td>'
            + '<td class="ca">' + amount + '</td>'
            + '<td class="ci">' + image_html + '</td>'
            + '</tr>'
        )

    total_qty = period_data["drawing_qty"].fillna(0).sum()
    total_amount = period_data["amount"].fillna(0).sum()
    total_label = "TOTAL IN " + str(period_data.iloc[0].get("period_label") or "").upper()
    rows_parts.append(
        '<tr class="tr">'
        + '<td class="cn"></td>'
        + '<td class="cp"><div class="pn">' + escape(total_label) + '</div></td>'
        + '<td class="cq">' + f"{total_qty:g}" + '</td>'
        + '<td class="cu"></td>'
        + '<td class="ca">SGD ' + f"{total_amount:,.0f}" + '</td>'
        + '<td class="ci"></td>'
        + '</tr>'
    )
    rows_html = "".join(rows_parts)

    est_height = max(420, 72 + estimated_body_height + 96)

    html = (
        '<!DOCTYPE html><html><head><meta charset="utf-8"><style>'
        + ':root{--bg:#FFFFFF;--sf:#F7F7F5;--bd:#E8E8E4;--tx:#1A1A1A;--mt:#888884;--ac:#B8760A;--pr:#2D2D2D;--sa:#F6E8D0;}'
        + 'body{margin:0;padding:4px 0;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;background:var(--bg);}'
        + '#' + wrap_id + '{border:1px solid var(--bd);border-radius:12px;overflow:auto;background:var(--bg);}'
        + '.pt{width:100%;min-width:900px;border-collapse:collapse;table-layout:fixed;font-size:14px;}'
        + '.pt th{position:sticky;top:0;z-index:10;background:var(--pr);color:white;padding:12px 14px;text-align:left;font-size:13px;font-weight:600;}'
        + '.pt td{padding:12px 14px;vertical-align:top;border-bottom:1px solid var(--bd);line-height:1.4;overflow-wrap:anywhere;}'
        + '.cn{width:46px;text-align:center;color:var(--mt);}'
        + '.cp{width:42%;}'
        + '.pn{font-size:15px;font-weight:700;color:var(--tx);}'
        + '.poi{color:var(--mt);font-size:13px;font-weight:500;}'
        + '.pd{margin-top:6px;}'
        + '.pdl{display:flex;align-items:flex-start;gap:6px;font-size:13px;color:var(--mt);margin-top:3px;}'
        + '.dot{width:4px;height:4px;margin-top:7px;border-radius:50%;background:var(--mt);flex:0 0 auto;}'
        + '.ci{width:290px;}'
        + '.image-grid{display:grid;grid-template-columns:repeat(3,84px);gap:6px;}'
        + '.image-cell{width:84px;aspect-ratio:1;border:1px solid var(--bd);border-radius:6px;overflow:hidden;display:flex;align-items:center;justify-content:center;padding:0;background:var(--sf);cursor:pointer;}'
        + '.image-cell:hover{border-color:var(--ac);box-shadow:0 0 0 2px var(--sa);}'
        + '.image-cell img{width:100%;height:100%;object-fit:contain;}'
        + '.cq{width:100px;text-align:right;}'
        + '.cu{width:120px;text-align:right;color:var(--mt);}'
        + '.ca{width:130px;text-align:right;font-weight:700;color:var(--ac);background:var(--sa);border-left:1px solid var(--bd);}'
        + '.tr td{background:var(--sf)!important;font-weight:700;border-top:2px solid var(--bd);}'
        + '.tr .ca{background:var(--sa)!important;}'
        + '#modal_' + wrap_id + '{display:none;position:fixed;inset:0;z-index:99999;background:rgba(0,0,0,.88);align-items:center;justify-content:center;padding:40px;}'
        + '#modal_' + wrap_id + '.open{display:flex;}'
        + '#modal_' + wrap_id + ' img{max-width:min(96vw,1400px);max-height:90vh;object-fit:contain;border-radius:8px;box-shadow:0 20px 80px rgba(0,0,0,.6);}'
        + '#mcls_' + wrap_id + '{position:absolute;top:20px;right:20px;width:44px;height:44px;border:0;border-radius:50%;background:rgba(255,255,255,.2);color:#fff;font-size:30px;cursor:pointer;display:flex;align-items:center;justify-content:center;}'
        + '</style></head><body>'
        + '<div id="' + wrap_id + '">'
        + '<table class="pt"><thead><tr>'
        + '<th class="cn">STT</th>'
        + '<th class="cp">Dự án</th>'
        + '<th class="cq">Số lượng</th>'
        + '<th class="cu">Đơn giá (SGD)</th>'
        + '<th class="ca">Tổng tiền (SGD)</th>'
        + '<th class="ci">Hình ảnh</th>'
        + '</tr></thead><tbody>'
        + rows_html
        + '</tbody></table></div>'
        + '<div id="modal_' + wrap_id + '"><button id="mcls_' + wrap_id + '">&#215;</button><img id="mimg_' + wrap_id + '" src="" alt="full"/></div>'
        + '<script>'
        + '(function(){'
        + 'var mo=document.getElementById("modal_' + wrap_id + '");'
        + 'var mi=document.getElementById("mimg_' + wrap_id + '");'
        + 'var cl=document.getElementById("mcls_' + wrap_id + '");'
        + 'var tbl=document.getElementById("' + wrap_id + '");'
        + 'if(!mo||!mi||!cl||!tbl)return;'
        + 'function close(){mo.classList.remove("open");setTimeout(function(){mi.src="";},300);}'
        + 'cl.addEventListener("click",close);'
        + 'mo.addEventListener("click",function(e){if(e.target===mo)close();});'
        + 'document.addEventListener("keydown",function(e){if(e.key==="Escape")close();});'
        + 'tbl.addEventListener("click",function(e){'
        + '  var t=e.target.closest(".image-open");if(!t)return;'
        + '  var s=t.getAttribute("data-full");if(!s)return;'
        + '  e.preventDefault();e.stopPropagation();'
        + '  mi.src=s;mo.classList.add("open");'
        + '},true);'
        + 'function resizeFrame(){'
        + '  var h=Math.ceil(document.documentElement.scrollHeight)+12;'
        + '  if(window.frameElement){window.frameElement.style.height=h+"px";}'
        + '}'
        + 'resizeFrame();setTimeout(resizeFrame,100);setTimeout(resizeFrame,500);'
        + 'if(window.ResizeObserver){new ResizeObserver(resizeFrame).observe(document.body);}'
        + '})();'
        + '</script>'
        + '</body></html>'
    )

    import streamlit.components.v1 as components
    components.html(html, height=est_height, scrolling=False)


def make_pdf(data: pd.DataFrame, title: str, include_period: bool = False, include_images: bool = True) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    body = ParagraphStyle("BodySmall", parent=styles["BodyText"], fontSize=8, leading=10)
    heading = ParagraphStyle("Heading", parent=styles["Title"], fontSize=16, leading=20, textColor=colors.HexColor("#111827"))
    project_style = ParagraphStyle("ProjectCell", parent=body, fontSize=8, leading=10)
    payment_style = ParagraphStyle("PaymentInfo", parent=styles["BodyText"], fontSize=9, leading=12)

    story = [Paragraph(title, heading), Spacer(1, 5 * mm)]
    total_qty = data["drawing_qty"].fillna(0).sum() if not data.empty else 0
    total_amount = data["amount"].fillna(0).sum() if not data.empty else 0
    story.append(Paragraph(f"Total drawings: {total_qty:g} | Total amount: SGD {total_amount:,.2f}", styles["Normal"]))
    story.append(Spacer(1, 5 * mm))

    if include_period:
        table_rows = [["No", "Period", "Project", "Qty", "Unit", "Amount"] + (["Image"] if include_images else [])]
    else:
        table_rows = [["No", "Project", "Qty", "Unit", "Amount"] + (["Image"] if include_images else [])]
    row_heights = [9 * mm]
    sorted_data = data.sort_values(["source_file", "source_row", "id"], na_position="last").reset_index(drop=True)
    for index, row in sorted_data.iterrows():
        project_title, owner_line, description_lines, _started_text = split_period_project_text(
            row.get("project_name"),
            row.get("owner"),
            row.get("description"),
        )
        heading_text = project_title
        if owner_line:
            heading_text = f"{heading_text} · {owner_line}"
        project_parts = [f"<b>{escape(heading_text)}</b>"]
        if description_lines:
            project_parts.extend(f"<i>{escape(line)}</i>" for line in description_lines)
        project_html = "<br/>".join(project_parts)
        project_text_for_height = "\n".join([heading_text] + description_lines)
        image_grid = pdf_image_grid(row.get("image_path")) if include_images else ""
        row_cells = [
                str(index + 1),
            ]
        if include_period:
            row_cells.append(Paragraph(escape(str(row.get("period_label") or "")).replace("\n", "<br/>"), project_style))
        row_cells.extend(
            [
                Paragraph(project_html, project_style),
                f"{number_or_zero(row.get('drawing_qty')):g}",
                f"SGD {number_or_zero(row.get('unit_price')):,.0f}",
                f"SGD {number_or_zero(row.get('amount')):,.0f}",
            ]
        )
        if include_images:
            row_cells.append(image_grid)
        table_rows.append(row_cells)
        height_lines = [project_text_for_height]
        if include_period:
            height_lines.append(str(row.get("period_label") or ""))
        row_heights.append(calc_pdf_row_height(row.get("image_path") if include_images else "", height_lines, include_period=include_period))
    if include_period:
        total_row = ["", "", "TOTAL", f"{total_qty:g}", "", f"SGD {total_amount:,.0f}"]
    else:
        total_row = ["", "TOTAL", f"{total_qty:g}", "", f"SGD {total_amount:,.0f}"]
    if include_images:
        total_row.append("")
    table_rows.append(total_row)
    row_heights.append(9 * mm)

    if include_period and include_images:
        col_widths = [8 * mm, 32 * mm, 52 * mm, 12 * mm, 20 * mm, 24 * mm, 36 * mm]
    elif include_period:
        col_widths = [9 * mm, 38 * mm, 82 * mm, 14 * mm, 24 * mm, 28 * mm]
    elif include_images:
        col_widths = [10 * mm, 70 * mm, 14 * mm, 23 * mm, 28 * mm, 50 * mm]
    else:
        col_widths = [12 * mm, 100 * mm, 18 * mm, 28 * mm, 34 * mm]
    amount_col = 5 if include_period else 4
    image_col = (6 if include_period else 5) if include_images else None
    table = Table(table_rows, colWidths=col_widths, rowHeights=row_heights, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E5E7EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9D9D9")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 1), (0, -1), "CENTER"),
                ("ALIGN", (3 if include_period else 2, 1), (amount_col, -1), "RIGHT"),
                ("BACKGROUND", (amount_col, 1), (amount_col, -1), colors.HexColor("#FFF7ED")),
                ("TEXTCOLOR", (amount_col, 1), (amount_col, -1), colors.HexColor("#B45309")),
                ("FONTNAME", (amount_col, 1), (amount_col, -1), "Helvetica-Bold"),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E5E7EB")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ]
        )
    )
    if image_col is not None:
        table.setStyle(TableStyle([("ALIGN", (image_col, 1), (image_col, -1), "CENTER")]))
    story.append(table)
    payment_info = get_setting("payment_info", DEFAULT_PAYMENT_INFO)
    if payment_info.strip():
        story.append(Spacer(1, 8 * mm))
        story.append(Paragraph("Payment information", styles["Heading3"]))
        story.append(Paragraph(escape(payment_info).replace("\n", "<br/>"), payment_style))
    doc.build(story)
    return buffer.getvalue()


def save_pdf_file(data: pd.DataFrame, title: str, include_period: bool = False, include_images: bool = True) -> Path:
    date_str = datetime.now().strftime("%Y-%m-%d")
    target_dir = EXPORT_DIR / date_str
    target_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = target_dir / f"{safe_filename(title)}_{timestamp}.pdf"
    output_path.write_bytes(make_pdf(data, title, include_period=include_period, include_images=include_images))
    return output_path


def save_excel_file(data: pd.DataFrame, title: str, include_period: bool = False) -> Path:
    date_str = datetime.now().strftime("%Y-%m-%d")
    target_dir = EXPORT_DIR / date_str
    target_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = target_dir / f"{safe_filename(title)}_{timestamp}.xlsx"
    # ... rest of function logic ...

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Period"

    headers = ["No", "Period", "Project", "Drawings Quantity", "Unit Price (SGD)", "Amount (SGD)", "Image"] if include_period else ["No", "Project", "Drawings Quantity", "Unit Price (SGD)", "Amount (SGD)", "Image"]
    image_start_col = 7 if include_period else 6
    image_end_col = image_start_col + 2
    project_col = 3 if include_period else 2
    qty_col = 4 if include_period else 3
    unit_col = 5 if include_period else 4
    amount_col = 6 if include_period else 5
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    amount_fill = PatternFill("solid", fgColor="FFF7ED")
    thin_side = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    strong_border = Border(top=Side(style="medium", color="9CA3AF"))

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=image_end_col)
    title_cell = sheet.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=16, color="111827")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 26

    total_qty = data["drawing_qty"].fillna(0).sum() if not data.empty else 0
    total_amount = data["amount"].fillna(0).sum() if not data.empty else 0
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=image_end_col)
    summary_cell = sheet.cell(row=2, column=1, value=f"Total drawings: {total_qty:g} | Total amount: SGD {total_amount:,.2f}")
    summary_cell.alignment = Alignment(horizontal="center", vertical="center")
    summary_cell.font = Font(color="374151")

    header_row = 4
    for col_index, header in enumerate(headers[:-1], start=1):
        cell = sheet.cell(row=header_row, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="111827")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    sheet.merge_cells(start_row=header_row, start_column=image_start_col, end_row=header_row, end_column=image_end_col)
    image_header = sheet.cell(row=header_row, column=image_start_col, value="Image")
    image_header.fill = header_fill
    image_header.font = Font(bold=True, color="111827")
    image_header.alignment = Alignment(horizontal="center", vertical="center")
    for col_index in range(image_start_col, image_end_col + 1):
        sheet.cell(header_row, col_index).border = border

    widths = (
        {1: 6, 2: 28, 3: 44, 4: 18, 5: 18, 6: 18, 7: 15, 8: 15, 9: 15}
        if include_period
        else {1: 6, 2: 48, 3: 18, 4: 18, 5: 18, 6: 15, 7: 15, 8: 15}
    )
    for col_index, width in widths.items():
        sheet.column_dimensions[get_column_letter(col_index)].width = width

    sorted_data = data.sort_values(["source_file", "source_row", "id"], na_position="last").reset_index(drop=True)
    for index, row in sorted_data.iterrows():
        excel_row = header_row + 1 + index
        project_title, owner_line, description_lines, _started_text = split_period_project_text(
            row.get("project_name"),
            row.get("owner"),
            row.get("description"),
        )
        project_heading = project_title
        if owner_line:
            project_heading = f"{project_heading} · {owner_line}"
        description = "\n".join(description_lines)
        project_cell_value: object
        if description:
            project_cell_value = CellRichText(
                TextBlock(InlineFont(b=True), project_heading),
                TextBlock(InlineFont(i=True), f"\n{description}"),
            )
        else:
            project_cell_value = CellRichText(TextBlock(InlineFont(b=True), project_heading))

        sheet.cell(excel_row, 1, index + 1)
        if include_period:
            sheet.cell(excel_row, 2, str(row.get("period_label") or ""))
        sheet.cell(excel_row, project_col, project_cell_value)
        sheet.cell(excel_row, qty_col, number_or_zero(row.get("drawing_qty")))
        sheet.cell(excel_row, unit_col, number_or_zero(row.get("unit_price")))
        sheet.cell(excel_row, amount_col, number_or_zero(row.get("amount")))
        sheet.cell(excel_row, unit_col).number_format = '"SGD" #,##0'
        sheet.cell(excel_row, amount_col).number_format = '"SGD" #,##0'
        sheet.cell(excel_row, amount_col).fill = amount_fill
        sheet.cell(excel_row, amount_col).font = Font(bold=True, color="B45309")
        for col_index in range(1, image_end_col + 1):
            sheet.cell(excel_row, col_index).alignment = Alignment(vertical="top", wrap_text=True)
            sheet.cell(excel_row, col_index).border = border
        for col_index in [1, qty_col, unit_col, amount_col]:
            sheet.cell(excel_row, col_index).alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

        image_paths = image_paths_from_value(row.get("image_path"))[:9]
        image_count = 0
        for image_index, path_text in enumerate(image_paths):
            path = resolve_image_path(path_text)
            if not path.exists():
                continue
            image = XLImage(str(path))
            image.width = 72
            image.height = 72
            col_offset = image_count % 3
            image.anchor = f"{get_column_letter(image_start_col + col_offset)}{excel_row}"
            sheet.add_image(image)
            image_count += 1
        rows_needed = max(1, (image_count + 2) // 3)
        sheet.row_dimensions[excel_row].height = max(54, rows_needed * 56)

    total_row = header_row + len(sorted_data) + 1
    sheet.cell(total_row, project_col, "TOTAL")
    sheet.cell(total_row, qty_col, total_qty)
    sheet.cell(total_row, amount_col, total_amount)
    sheet.cell(total_row, amount_col).number_format = '"SGD" #,##0'
    for col_index in range(1, image_end_col + 1):
        cell = sheet.cell(total_row, col_index)
        cell.border = strong_border
        cell.font = Font(bold=True)
    sheet.cell(total_row, amount_col).font = Font(bold=True, color="B45309")

    payment_info = get_setting("payment_info", DEFAULT_PAYMENT_INFO)
    if payment_info.strip():
        payment_title_row = total_row + 2
        payment_row = total_row + 3
        sheet.merge_cells(start_row=payment_title_row, start_column=1, end_row=payment_title_row, end_column=image_end_col)
        sheet.cell(payment_title_row, 1, "Payment information")
        sheet.cell(payment_title_row, 1).font = Font(bold=True, size=12, color="111827")
        sheet.merge_cells(start_row=payment_row, start_column=1, end_row=payment_row, end_column=image_end_col)
        payment_cell = sheet.cell(payment_row, 1, payment_info)
        payment_cell.alignment = Alignment(vertical="top", wrap_text=True)
        payment_cell.border = border
        visual_lines = sum(max(1, (len(line) // 95) + 1) for line in payment_info.splitlines())
        sheet.row_dimensions[payment_row].height = max(95, visual_lines * 18)

    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.45
    sheet.page_margins.bottom = 0.45
    sheet.print_title_rows = f"{header_row}:{header_row}"
    workbook.save(output_path)
    return output_path


def save_uploaded_images(uploaded_files: list[object]) -> list[str]:
    ORIGINAL_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    THUMB_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    if "saved_uploads" not in st.session_state:
        st.session_state.saved_uploads = {}
    saved_paths = []
    for uploaded_file in uploaded_files:
        file_bytes = uploaded_file.getvalue()
        upload_key = hashlib.sha256(file_bytes).hexdigest()
        if upload_key in st.session_state.saved_uploads:
            saved_paths.append(st.session_state.saved_uploads[upload_key])
            continue
        stem = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        original_suffix = Path(uploaded_file.name).suffix.lower()
        if original_suffix not in {".png", ".jpg", ".jpeg", ".webp"}:
            original_suffix = ".jpg"
        original_path = ORIGINAL_UPLOAD_DIR / f"{stem}{original_suffix}"
        thumb_path = THUMB_UPLOAD_DIR / f"{stem}.jpg"
        with Image.open(BytesIO(file_bytes)) as image:
            if original_suffix in {".jpg", ".jpeg"}:
                image.convert("RGB").save(original_path, format="JPEG", quality=95)
            elif original_suffix == ".png":
                image.save(original_path, format="PNG")
            elif original_suffix == ".webp":
                image.save(original_path, format="WEBP", quality=95)
            else:
                image.convert("RGB").save(original_path, format="JPEG", quality=95)
            thumb = image.convert("RGB")
            thumb.thumbnail((320, 320))
            thumb.save(thumb_path, format="JPEG", quality=82, optimize=True)
        st.session_state.saved_uploads[upload_key] = str(thumb_path)
        saved_paths.append(str(thumb_path))
    return saved_paths


def display_metrics(data: pd.DataFrame) -> None:
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Dự án", f"{data['project_name'].nunique() if not data.empty else 0}")
    col2.metric("Dòng", f"{len(data)}")
    col3.metric("Drawings", f"{data['drawing_qty'].fillna(0).sum():g}" if not data.empty else "0")
    col4.metric("Tổng tiền SGD", f"{data['amount'].fillna(0).sum():,.0f}" if not data.empty else "0")


def inject_app_css() -> None:
    st.markdown(
        """
        <style>
        :root {
            --bg: #FFFFFF;
            --surface: #F7F7F5;
            --border: #E8E8E4;
            --text: #1A1A1A;
            --muted: #888884;
            --accent: #B8760A;
            --primary: #2D2D2D;
            --soft-accent: #F6E8D0;
        }
        html, body, [class*="css"] {
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
        }
        .stApp {
            background: var(--surface) !important;
            color: var(--text);
        }
        div[data-testid="stAppViewContainer"],
        div[data-testid="stMain"],
        div[data-testid="stMainBlockContainer"] {
            background: var(--surface) !important;
            color: var(--text);
        }
        section[data-testid="stSidebar"] {
            background: var(--bg) !important;
            border-right: 1px solid var(--border);
        }
        section[data-testid="stSidebar"] > div {
            padding-top: 1.4rem;
        }
        .block-container {
            padding-top: 2rem;
            padding-bottom: 2rem;
            max-width: 1500px;
        }
        h1, h2, h3 {
            letter-spacing: 0;
            color: var(--text);
        }
        h1 {
            font-size: 22px;
            font-weight: 600;
        }
        h2 {
            font-size: 18px;
            font-weight: 600;
        }
        h3 {
            font-size: 15px;
            font-weight: 500;
        }
        div[data-testid="stMetric"] {
            background: var(--bg);
            border: 1px solid var(--border);
            border-radius: 16px;
            padding: 14px 16px;
        }
        div[data-testid="stMetric"] label {
            color: var(--muted);
        }
        .bento-grid {
            display: grid;
            grid-template-columns: repeat(4, minmax(0, 1fr));
            gap: 16px;
            padding: 8px 0 16px;
        }
        .bento-sm { grid-column: span 1; }
        .bento-md { grid-column: span 2; }
        .bento-lg { grid-column: span 3; }
        .bento-full { grid-column: span 4; }
        @media (max-width: 900px) {
            .bento-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
            .bento-lg, .bento-full { grid-column: span 2; }
        }
        @media (max-width: 620px) {
            .bento-grid { grid-template-columns: 1fr; }
            .bento-sm, .bento-md, .bento-lg, .bento-full { grid-column: span 1; }
        }
        .bento-card {
            background: var(--bg);
            border: 1px solid var(--border);
            border-radius: 16px;
            padding: 20px 24px;
            min-height: 120px;
        }
        .bento-title {
            font-size: 12px;
            font-weight: 500;
            color: var(--muted);
            text-transform: uppercase;
            letter-spacing: 0.06em;
            margin-bottom: 8px;
        }
        .bento-value {
            font-size: 32px;
            font-weight: 600;
            color: var(--primary);
            line-height: 1;
        }
        .bento-note {
            font-size: 13px;
            color: var(--muted);
            margin-top: 6px;
        }
        .client-row, .activity-row {
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 12px;
            padding: 8px 0;
            border-bottom: 1px solid var(--border);
            font-size: 0.92rem;
        }
        .client-row:last-child, .activity-row:last-child {
            border-bottom: 0;
        }
        .client-left {
            display: flex;
            align-items: center;
            gap: 8px;
            min-width: 0;
        }
        .client-dot {
            width: 8px;
            height: 8px;
            border-radius: 999px;
            display: inline-block;
            flex: 0 0 auto;
            background: var(--primary);
        }
        .tone-0 { --tone: var(--accent); }
        .tone-1 { --tone: #6F6F6A; }
        .tone-2 { --tone: #A67C52; }
        .tone-3 { --tone: #4D5A63; }
        .tone-4 { --tone: #8A8A84; }
        .tone-5 { --tone: #C08A2C; }
        .tone-0 .client-dot, .tone-1 .client-dot, .tone-2 .client-dot,
        .tone-3 .client-dot, .tone-4 .client-dot, .tone-5 .client-dot,
        .tone-0 .mini-bar > span, .tone-1 .mini-bar > span, .tone-2 .mini-bar > span,
        .tone-3 .mini-bar > span, .tone-4 .mini-bar > span, .tone-5 .mini-bar > span {
            background: var(--tone);
        }
        .row-right {
            text-align: right;
            color: var(--text);
        }
        .row-subtle {
            font-size: 0.78rem;
            color: var(--muted);
        }
        .mini-bar {
            width: 100%;
            height: 8px;
            background: var(--border);
            border-radius: 999px;
            overflow: hidden;
            margin-top: 4px;
            border: 1px solid var(--border);
        }
        .mini-bar > span {
            display: block;
            height: 100%;
            background: var(--primary);
        }
        .dashboard-chart {
            display: flex;
            align-items: end;
            gap: 14px;
            height: 320px;
            padding-top: 8px;
            padding-bottom: 4px;
            border-bottom: 1px solid var(--border);
            overflow-x: auto;
        }
        .chart-toggle-btn {
            padding: 3px 10px; font-size: 12px; border: 1px solid var(--border);
            border-radius: 6px; background: var(--bg); color: var(--muted);
            cursor: pointer; transition: all .15s;
        }
        .chart-toggle-btn.active { background: var(--primary); color: white; border-color: var(--primary); }
        .chart-col {
            flex: 1 0 64px;
            min-width: 64px;
            display: flex;
            flex-direction: column;
            justify-content: end;
            align-items: center;
            gap: 6px;
            height: 100%;
        }
        .chart-value {
            font-size: 11px;
            color: var(--accent);
            font-weight: 600;
            white-space: nowrap;
        }
        .chart-bar {
            width: 100%;
            max-width: 44px;
            min-height: 4px;
            border-radius: 8px 8px 0 0;
            background: var(--accent);
        }
        .chart-label {
            font-size: 11px;
            color: var(--muted);
            text-align: center;
            line-height: 1.3;
            white-space: normal;
            overflow-wrap: break-word;
            max-width: 90px;
        }
        .chart-company {
            font-size: 11px;
            font-weight: 600;
            text-align: center;
            line-height: 1.3;
            white-space: normal;
            overflow-wrap: break-word;
            max-width: 90px;
        }
        .top-project-row {
            display: grid;
            grid-template-columns: 1fr auto;
            gap: 10px;
            padding: 9px 0;
            border-bottom: 1px solid var(--border);
            font-size: 13px;
        }
        .top-project-row:last-child {
            border-bottom: 0;
        }
        .top-project-name {
            color: var(--text);
            font-weight: 500;
            overflow-wrap: anywhere;
        }
        .top-project-amount {
            color: var(--accent);
            font-weight: 600;
            white-space: nowrap;
        }
        .summary-strip {
            background: var(--bg);
            border: 1px solid var(--border);
            border-radius: 12px;
            padding: 10px 12px;
            color: var(--muted);
            font-size: 0.92rem;
        }
        .upload-grid {
            display: grid;
            grid-template-columns: repeat(3, 56px);
            gap: 7px;
            margin-top: 8px;
        }
        .upload-thumb {
            width: 56px;
            aspect-ratio: 1;
            border: 1px solid var(--border);
            border-radius: 6px;
            overflow: hidden;
            display: flex;
            align-items: center;
            justify-content: center;
            background: var(--bg);
            color: var(--muted);
            font-size: 12px;
            font-weight: 650;
            cursor: pointer;
        }
        .upload-thumb img {
            width: 100%;
            height: 100%;
            object-fit: contain;
        }
        .upload-more {
            background: var(--surface);
        }
        .gallery-grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(160px, 1fr));
            gap: 14px;
            margin-top: 8px;
        }
        .gallery-card {
            border: 1px solid var(--border);
            border-radius: 10px;
            overflow: hidden;
            background: var(--bg);
            display: flex;
            flex-direction: column;
        }
        .gallery-thumb {
            width: 100%;
            aspect-ratio: 1;
            border: 1px solid var(--border);
            border-radius: 10px;
            background: var(--surface);
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 0;
            overflow: hidden;
            transition: border-color 0.15s, box-shadow 0.15s;
        }
        .gallery-thumb:hover {
            border-color: var(--accent);
            box-shadow: 0 0 0 2px var(--soft-accent);
        }
        .gallery-thumb img {
            width: 100%;
            height: 100%;
            object-fit: contain;
        }
        .gallery-caption {
            padding: 8px 10px;
            font-size: 12px;
            line-height: 1.4;
        }
        .gallery-caption .gc-project {
            font-weight: 700;
            color: var(--text);
            overflow-wrap: anywhere;
        }
        .gallery-caption .gc-meta {
            color: var(--muted);
            margin-top: 2px;
        }
        .filter-card {
            background: var(--bg);
            border: 1px solid var(--border);
            border-radius: 12px;
            padding: 12px 14px 6px;
            margin-bottom: 12px;
        }
        .stTabs [data-baseweb="tab-list"] {
            gap: 8px;
        }
        .stTabs [data-baseweb="tab"] {
            background: transparent;
            border: 0;
            border-radius: 0;
            padding: 8px 14px;
            color: var(--muted);
            font-size: 14px;
            font-weight: 500;
        }
        .stTabs [aria-selected="true"] {
            color: var(--text);
            border-bottom: 2px solid var(--text);
        }
        .stTextInput input, .stTextArea textarea, .stSelectbox div[data-baseweb="select"] > div,
        .stNumberInput input {
            background-color: var(--bg);
            color: var(--text);
            border-color: var(--border);
        }
        div[data-testid="stVerticalBlockBorderWrapper"] {
            background: var(--bg);
            border-color: var(--border);
        }
        button,
        [role="button"],
        .stButton > button,
        select,
        [data-testid="stSelectbox"] > div,
        [data-testid="stSelectbox"] *,
        div[data-baseweb="select"],
        div[data-baseweb="select"] *,
        div[data-baseweb="select"] svg,
        [data-testid="baseButton-secondary"],
        [data-testid="baseButton-primary"],
        .sort-arrow,
        .clickable {
            cursor: pointer !important;
            user-select: none !important;
        }
        input,
        textarea,
        [contenteditable="true"],
        [data-baseweb="base-input"] input,
        [data-baseweb="textarea"] textarea {
            cursor: text !important;
            user-select: text !important;
        }
        div[data-baseweb="select"] input {
            cursor: pointer !important;
            caret-color: transparent !important;
            user-select: none !important;
        }

        /* Global Image Modal */
        .image-open { cursor: pointer; }
        .image-modal {
            position: fixed;
            inset: 0;
            z-index: 10000001;
            display: none;
            align-items: center;
            justify-content: center;
            padding: 40px;
            background: rgba(240,240,238,.93);
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
        }
        .image-modal.is-open { display: flex; }
        .image-modal img {
            max-width: min(96vw, 1400px);
            max-height: 92vh;
            object-fit: contain;
            border-radius: 8px;
            box-shadow: 0 20px 80px rgba(0, 0, 0, 0.6);
            background: #000;
        }
        .image-modal-close {
            position: absolute;
            top: 24px;
            right: 24px;
            width: 48px;
            height: 48px;
            border: 0;
            border-radius: 50%;
            background: rgba(255, 255, 255, 0.15);
            color: white;
            font-size: 36px;
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-content: center;
            transition: all 0.2s;
        }
        .image-modal-close:hover {
            background: rgba(255, 255, 255, 0.3);
            transform: scale(1.1);
        }

        /* Settings tab pills */
        .stTabs [data-baseweb="tab-list"] {
            gap: 4px;
            background: var(--surface);
            padding: 4px;
            border-radius: 10px;
        }
        .stTabs [data-baseweb="tab"] {
            padding: 8px 14px;
            border-radius: 8px;
            font-size: 13px;
            font-weight: 500;
            color: var(--muted);
            background: transparent;
            border: 0;
        }
        .stTabs [aria-selected="true"] {
            background: var(--bg);
            color: var(--text);
            box-shadow: 0 1px 3px rgba(0,0,0,0.06);
        }
        </style>

        <div id="globalImageModal" class="image-modal" aria-hidden="true">
            <button type="button" class="image-modal-close" aria-label="Close">×</button>
            <img id="globalImageModalImg" src="" alt="Full size" />
        </div>

        <script>
        (() => {
            const initModal = () => {
                if (window.imageModalInitialized) return;
                window.imageModalInitialized = true;

                // Event delegation on document to catch all clicks on .image-open
                document.addEventListener("click", (e) => {
                    const trigger = e.target.closest(".image-open");
                    if (trigger) {
                        const modal = document.getElementById("globalImageModal");
                        const modalImg = document.getElementById("globalImageModalImg");
                        if (!modal || !modalImg) return;

                        const fullSrc = trigger.getAttribute("data-full");
                        if (fullSrc) {
                            modalImg.src = fullSrc;
                            modal.classList.add("is-open");
                            modal.setAttribute("aria-hidden", "false");
                            e.preventDefault();
                            e.stopPropagation();
                        }
                    }
                }, true);

                // Handle closing (click on close btn or background)
                document.addEventListener("click", (e) => {
                    const modal = document.getElementById("globalImageModal");
                    if (!modal || !modal.classList.contains("is-open")) return;

                    if (e.target.closest(".image-modal-close") || e.target === modal) {
                        modal.classList.remove("is-open");
                        modal.setAttribute("aria-hidden", "true");
                        const modalImg = document.getElementById("globalImageModalImg");
                        if (modalImg) setTimeout(() => { modalImg.src = ""; }, 300);
                    }
                }, true);

                // Handle ESC key
                document.addEventListener("keydown", (e) => {
                    if (e.key === "Escape") {
                        const modal = document.getElementById("globalImageModal");
                        if (modal && modal.classList.contains("is-open")) {
                            modal.classList.remove("is-open");
                            modal.setAttribute("aria-hidden", "true");
                            const modalImg = document.getElementById("globalImageModalImg");
                            if (modalImg) setTimeout(() => { modalImg.src = ""; }, 300);
                        }
                    }
                }, true);
            };

            // Run initialization
            if (document.readyState === "loading") {
                document.addEventListener("DOMContentLoaded", initModal);
            } else {
                initModal();
            }
            
            // Re-check periodically in case Streamlit re-renders the root container
            setInterval(initModal, 1000);
        })();
        </script>
        """,
        unsafe_allow_html=True,
    )



def bento_card(title: str, value: str, note: str = "") -> None:
    st.markdown(
        f"""
        <div class="bento-card">
            <div class="bento-title">{escape(title)}</div>
            <div class="bento-value">{escape(value)}</div>
            <div class="bento-note">{escape(note)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def latest_period_label(data: pd.DataFrame) -> str:
    if data.empty:
        return "-"
    scoped = data.dropna(subset=["period_year"]).copy()
    if scoped.empty:
        return "-"
    scoped["_period_sort"] = scoped["period_label"].fillna("").map(period_sort_key)
    latest = scoped.sort_values(["period_year", "_period_sort", "updated_at"], ascending=[False, False, False]).iloc[0]
    return f"{uppercase_label(latest.get('client_name')) or 'GENERAL'} - {latest.get('period_label') or '-'}"


def display_client_breakdown(data: pd.DataFrame) -> None:
    st.markdown('<div class="bento-title">Theo công ty</div>', unsafe_allow_html=True)
    if data.empty:
        st.caption("Chưa có dữ liệu.")
        return
    grouped = (
        data.assign(_client_upper=data["client_name"].map(uppercase_label)).groupby("_client_upper", dropna=False)
        .agg(rows=("id", "count"), amount=("amount", "sum"))
        .reset_index()
        .sort_values("amount", ascending=False)
    )
    max_amount = max(float(grouped["amount"].max() or 0), 1.0)
    rows_html = []
    for index, row in grouped.iterrows():
        amount = number_or_zero(row["amount"])
        width = max(4, min(100, int((amount / max_amount) * 100)))
        tone_class = f"tone-{len(rows_html) % 6}"
        rows_html.append(
            f"""
            <div class="client-row {tone_class}">
                <div style="flex:1; min-width:0;">
                    <div class="client-left">
                        <span class="client-dot"></span>
                        <strong>{escape(uppercase_label(row['_client_upper']) or 'GENERAL')}</strong>
                    </div>
                    <div class="mini-bar"><span style="width:{width}%;"></span></div>
                </div>
                <div class="row-right">
                    <div>SGD {amount:,.0f}</div>
                    <div class="row-subtle">{int(row['rows'])} dòng</div>
                </div>
            </div>
            """
        )
    st.markdown("".join(rows_html), unsafe_allow_html=True)


def display_recent_activity(data: pd.DataFrame) -> None:
    st.markdown('<div class="bento-title">Hoạt động gần đây</div>', unsafe_allow_html=True)
    if data.empty or "updated_at" not in data.columns:
        st.caption("Chưa có hoạt động.")
        return
    recent = data.sort_values("updated_at", ascending=False).head(5)
    rows_html = []
    for index, row in recent.iterrows():
        tone_class = f"tone-{index % 6}"
        rows_html.append(
            f"""
            <div class="activity-row {tone_class}">
                <div style="min-width:0;">
                    <div><span class="client-dot"></span>
                    <strong>{escape(uppercase_project_name(row.get('project_name')))}</strong></div>
                    <div class="row-subtle">{escape(uppercase_label(row.get('client_name')) or 'GENERAL')} · {escape(str(row.get('period_label') or ''))}</div>
                </div>
                <div class="row-subtle row-right">{escape(str(row.get('updated_at') or ''))}</div>
            </div>
            """
        )
    st.markdown("".join(rows_html), unsafe_allow_html=True)


# Palette: one color per company (up to 8)
_CLIENT_PALETTE = ["#B8760A","#4A7FA5","#5C9E6E","#9B5CA5","#C0694A","#4A8FA0","#A07840","#7A7A7A"]

def _client_color_map(data: pd.DataFrame) -> dict:
    clients = sorted(data["client_name"].map(uppercase_label).dropna().unique().tolist())
    return {c: _CLIENT_PALETTE[i % len(_CLIENT_PALETTE)] for i, c in enumerate(clients)}

def _build_chart_bars(grouped_data: pd.DataFrame, color_map: dict, max_amount: float) -> str:
    bars = []
    for _, row in grouped_data.iterrows():
        amount = number_or_zero(row["amount"])
        height = max(4, min(100, int(amount / max_amount * 100)))
        label = str(row.get("label") or "-")
        display_label = label.replace(" - ", " ").replace(" + ", " + ")
        client = str(row.get("client") or "GENERAL")
        color = color_map.get(client, "#888884")
        bars.append(
            f'<div class="chart-col" title="{escape(client)} · {escape(label)} · SGD {amount:,.0f}">'
            f'<div class="chart-value">SGD {amount:,.0f}</div>'
            f'<div class="chart-bar" style="height:{height}%;background:{color};"></div>'
            f'<div class="chart-label">{escape(display_label)}</div>'
            f'<div class="chart-company" style="color:{color};">{escape(client)}</div>'
            f'</div>'
        )
    return "".join(bars)


def _build_legend(clients: list, color_map: dict) -> str:
    return "".join(
        f'<span style="display:inline-flex;align-items:center;gap:5px;margin-right:14px;font-size:12px;color:var(--muted);">'
        f'<span style="width:10px;height:10px;border-radius:2px;background:{color_map.get(c, "#888")};flex:0 0 auto;"></span>'
        f'{escape(c)}</span>'
        for c in clients
    )


def revenue_donut_html(data: pd.DataFrame) -> str:
    if data.empty:
        return '<div class="row-subtle">Chưa có dữ liệu.</div>'
    color_map = _client_color_map(data)
    grouped = (
        data.assign(_client_upper=data["client_name"].map(uppercase_label).fillna("GENERAL"))
        .groupby("_client_upper", dropna=False)
        .agg(amount=("amount", "sum"))
        .reset_index()
        .sort_values("amount", ascending=False)
    )
    grouped = grouped[grouped["amount"] > 0]
    total = float(grouped["amount"].sum() or 0)
    if total <= 0:
        return '<div class="row-subtle">Chưa có dữ liệu.</div>'

    radius = 60
    circumference = 2 * 3.14159265358979 * radius
    offset = 0.0
    segments = []
    for _, row in grouped.iterrows():
        client = str(row["_client_upper"]) or "GENERAL"
        amount = float(row["amount"])
        fraction = amount / total
        dash = fraction * circumference
        color = color_map.get(client, "#888884")
        segments.append(
            f'<circle r="{radius}" cx="80" cy="80" fill="transparent" '
            f'stroke="{color}" stroke-width="24" '
            f'stroke-dasharray="{dash:.3f} {circumference - dash:.3f}" '
            f'stroke-dashoffset="{-offset:.3f}" '
            f'transform="rotate(-90 80 80)">'
            f'<title>{escape(client)}: SGD {amount:,.0f} ({fraction * 100:.1f}%)</title>'
            f'</circle>'
        )
        offset += dash

    svg = (
        '<svg viewBox="0 0 160 160" width="180" height="180" '
        'style="display:block;margin:0 auto;">'
        + "".join(segments)
        + f'<text x="80" y="76" text-anchor="middle" font-size="13" fill="var(--muted)">Tổng doanh thu</text>'
        + f'<text x="80" y="96" text-anchor="middle" font-size="15" font-weight="700" fill="var(--text)">SGD {total:,.0f}</text>'
        + '</svg>'
    )

    legend_rows = []
    for _, row in grouped.iterrows():
        client = str(row["_client_upper"]) or "GENERAL"
        amount = float(row["amount"])
        fraction = amount / total
        color = color_map.get(client, "#888884")
        legend_rows.append(
            '<div style="display:flex;align-items:center;justify-content:space-between;gap:10px;padding:5px 0;font-size:13px;">'
            f'<div style="display:flex;align-items:center;gap:8px;min-width:0;">'
            f'<span style="width:10px;height:10px;border-radius:2px;background:{color};flex:0 0 auto;"></span>'
            f'<span style="overflow-wrap:anywhere;">{escape(client)}</span>'
            '</div>'
            f'<div style="white-space:nowrap;color:var(--muted);">SGD {amount:,.0f} · {fraction * 100:.1f}%</div>'
            '</div>'
        )

    return (
        '<div style="display:flex;gap:20px;align-items:center;flex-wrap:wrap;">'
        f'<div style="flex:0 0 auto;">{svg}</div>'
        f'<div style="flex:1;min-width:200px;">{"".join(legend_rows)}</div>'
        '</div>'
    )


def _get_paginated_revenue_data(data: pd.DataFrame, page: int, page_size: int = 12):
    if data.empty:
        return data, 0, 0
    chart_data = data.copy()
    chart_data["_client_upper"] = chart_data["client_name"].map(uppercase_label)
    chart_data["_period_key"] = (
        chart_data["_client_upper"].fillna("GENERAL")
        + " · "
        + chart_data["period_label"].fillna("")
    )
    grouped = (
        chart_data.groupby("_period_key", dropna=False)
        .agg(
            amount=("amount", "sum"),
            year=("period_year", "max"),
            label=("period_label", "first"),
            client=("_client_upper", "first"),
        )
        .reset_index()
    )
    grouped["_sort"] = grouped["label"].fillna("").map(period_sort_key)
    grouped = grouped.sort_values(["year", "_sort"], ascending=[False, False])
    total_count = len(grouped)
    total_pages = (total_count - 1) // page_size + 1 if total_count else 0
    page = max(0, min(page, max(0, total_pages - 1)))
    page_data = grouped.iloc[page * page_size : (page + 1) * page_size].copy()
    page_data = page_data.sort_values(["year", "_sort"], ascending=[True, True])
    return page_data, total_pages, total_count


def display_revenue_chart_card(data: pd.DataFrame) -> None:
    if data.empty:
        st.markdown(
            '<div class="row-subtle" style="padding:20px;text-align:center;">Chưa có dữ liệu.</div>',
            unsafe_allow_html=True,
        )
        return
    color_map = _client_color_map(data)
    if "chart_page" not in st.session_state:
        st.session_state.chart_page = 0

    page_size = 8
    page_data, total_pages, total_count = _get_paginated_revenue_data(
        data, st.session_state.chart_page, page_size=page_size
    )
    if page_data.empty or total_pages == 0:
        st.markdown(
            '<div class="row-subtle" style="padding:20px;text-align:center;">Chưa có dữ liệu.</div>',
            unsafe_allow_html=True,
        )
        return

    max_amount = max(float(page_data["amount"].max() or 0), 1.0)
    clients_in_chart = page_data["client"].dropna().unique().tolist()
    legend_html = _build_legend(clients_in_chart, color_map)
    bars_html = _build_chart_bars(page_data, color_map, max_amount)
    oldest = page_data.iloc[0]["label"]
    newest = page_data.iloc[-1]["label"]
    range_info = f"{oldest} → {newest}"

    head_cols = st.columns([6, 1, 1, 1])
    with head_cols[0]:
        st.markdown(
            f'<div style="line-height:2;">{legend_html}</div>'
            f'<div style="font-size:11px;color:var(--muted);">{escape(range_info)}</div>',
            unsafe_allow_html=True,
        )
    with head_cols[1]:
        disabled_prev = st.session_state.chart_page <= 0
        if st.button("◀", key="chart_prev", disabled=disabled_prev, width="stretch",
                     help=f"Cũ hơn ({page_size} kỳ trước)"):
            st.session_state.chart_page -= 1
            st.rerun()
    with head_cols[2]:
        disabled_next = st.session_state.chart_page >= total_pages - 1
        if st.button("▶", key="chart_next", disabled=disabled_next, width="stretch",
                     help=f"Mới hơn ({page_size} kỳ sau)"):
            st.session_state.chart_page += 1
            st.rerun()
    with head_cols[3]:
        st.markdown(
            f'<div style="text-align:center;font-size:12px;color:var(--muted);padding-top:6px;">'
            f'{st.session_state.chart_page + 1}/{total_pages}</div>',
            unsafe_allow_html=True,
        )

    st.markdown(f'<div class="dashboard-chart" style="padding:8px 0;">{bars_html}</div>', unsafe_allow_html=True)


def top_projects_html(data: pd.DataFrame) -> str:
    if data.empty:
        return '<div class="row-subtle">Chưa có dữ liệu.</div>'
    scoped = data.copy()
    scoped["_project_upper"] = scoped["project_name"].map(uppercase_project_name)
    grouped = (
        scoped.groupby("_project_upper", dropna=False)
        .agg(amount=("amount", "sum"), rows=("id", "count"))
        .reset_index()
        .sort_values("amount", ascending=False)
        .head(5)
    )
    rows = []
    for _, row in grouped.iterrows():
        name = escape(str(row.get("_project_upper") or "-"))
        count = int(row.get("rows") or 0)
        amt = number_or_zero(row.get("amount"))
        rows.append(
            f'<div class="top-project-row"><div>'
            f'<div class="top-project-name">{name}</div>'
            f'<div class="row-subtle">{count} dòng</div>'
            f'</div><div class="top-project-amount">SGD {amt:,.0f}</div></div>'
        )
    return "".join(rows)


def display_bento_dashboard(data: pd.DataFrame) -> None:
    total_projects = data["project_name"].map(uppercase_project_name).nunique() if not data.empty else 0
    total_drawings = data["drawing_qty"].fillna(0).sum() if not data.empty else 0
    total_amount = data["amount"].fillna(0).sum() if not data.empty else 0
    clients_count = data["client_name"].map(uppercase_label).nunique() if not data.empty else 0
    latest_text = latest_period_label(data)
    latest_client, latest_period = (latest_text.split(" - ", 1) + ["-"])[:2] if " - " in latest_text else ("-", latest_text)

    _top_html = top_projects_html(data)
    _parts = [
        '<div class="bento-grid">',
        '<div class="bento-card bento-sm">',
        '<div class="bento-title">Tổng dự án</div>',
        f'<div class="bento-value">{total_projects:,}</div>',
        f'<div class="bento-note">{clients_count} công ty</div>',
        '</div>',
        '<div class="bento-card bento-sm">',
        '<div class="bento-title">Tổng drawings</div>',
        f'<div class="bento-value">{total_drawings:g}</div>',
        f'<div class="bento-note">{len(data):,} dòng dữ liệu</div>',
        '</div>',
        '<div class="bento-card bento-sm">',
        '<div class="bento-title">Doanh thu</div>',
        f'<div class="bento-value" style="color:var(--accent)">SGD {total_amount:,.0f}</div>',
        '<div class="bento-note">tổng active rows</div>',
        '</div>',
        '<div class="bento-card bento-sm">',
        '<div class="bento-title">Kỳ gần nhất</div>',
        f'<div class="bento-value" style="font-size:18px;line-height:1.2;">{escape(latest_period)}</div>',
        f'<div class="bento-note">{escape(latest_client)}</div>',
        '</div>',
        '</div>',
    ]
    st.markdown("".join(_parts), unsafe_allow_html=True)

    if data.empty:
        empty_state("Chưa có dữ liệu", icon="📦", subtitle="Vào Cài đặt > Import dữ liệu để bắt đầu.")
        return

    # Chart doanh thu theo kỳ (full width)
    with st.container(border=True):
        st.markdown("##### 📊 Doanh thu theo kỳ")
        display_revenue_chart_card(data)

    # Doanh thu theo công ty (donut) + Top projects
    pie_cols = st.columns(2, gap="medium")
    with pie_cols[0]:
        with st.container(border=True):
            st.markdown("##### 🥧 Doanh thu theo công ty")
            st.markdown(revenue_donut_html(data), unsafe_allow_html=True)
    with pie_cols[1]:
        with st.container(border=True):
            st.markdown("##### ⭐ Top projects")
            st.markdown(_top_html, unsafe_allow_html=True)

    # Company/period selector
    with st.container(border=True):
        st.markdown("##### 🔎 Xem nhanh theo công ty & kỳ")
        sel_cols = st.columns(2)
        with sel_cols[0]:
            selected_client = st.selectbox("Công ty", client_options(data), key="dashboard_client")
        scoped = data[data["client_name"].fillna("GENERAL").map(uppercase_label) == selected_client]
        period_options = periods_for_year(scoped, "All")
        with sel_cols[1]:
            selected_period = st.selectbox("Kỳ", period_options, key="dashboard_period")
        table_data = scoped[scoped["period_label"].fillna("") == selected_period]
        period_amount = table_data["amount"].fillna(0).sum()
        st.markdown(
            f'<div class="summary-strip">{escape(selected_client)} · {escape(selected_period)} · {len(table_data)} dòng · SGD {period_amount:,.0f}</div>',
            unsafe_allow_html=True,
        )
        display_search_table(table_data, key="dashboard_table")
        selected_id = st.session_state.get("dashboard_table_selected_id")
        if selected_id:
            display_edit_delete_panel(table_data, title="Sửa dòng đã chọn", selected_entry_id=int(selected_id))


def _jump_to_period_browser(client: str, period: str) -> None:
    """Set state so the 'Kỳ thanh toán' tab opens at the given client/period, then rerun."""
    # Use intermediate keys — NOT the widget keys (period_client, category_year)
    # display_period_browser will read these before rendering widgets
    st.session_state["_pending_period_client"] = client
    st.session_state["_pending_category_year"] = "All"
    st.session_state["pending_period_browser_period"] = period
    st.session_state["_jump_to_tab_label"] = "Kỳ thanh toán"
    st.rerun()


def display_period_browser(data: pd.DataFrame) -> None:
    if data.empty:
        empty_state("Chưa có dữ liệu", icon="📦", subtitle="Vào Cài đặt > Import dữ liệu để bắt đầu.")
        return

    # Apply pending navigation from gallery jump (must be done BEFORE widgets render)
    if "_pending_period_client" in st.session_state:
        pending_client = st.session_state.pop("_pending_period_client")
        pending_year = st.session_state.pop("_pending_category_year", "All")
        clients_all = client_options(data)
        if pending_client in clients_all:
            st.session_state["period_client"] = pending_client
        st.session_state["category_year"] = pending_year

    filter_cols = st.columns([1.2, 1, 1.7])
    clients = client_options(data)
    selected_client = filter_cols[0].selectbox("Công ty", clients, key="period_client")
    client_data = data[data["client_name"].fillna("GENERAL").map(uppercase_label) == selected_client]
    categories = year_options(client_data)
    if "All" not in categories:
        categories = ["All"] + categories
    selected_category = filter_cols[1].selectbox("Năm", categories, key="category_year")
    year_data = client_data if selected_category == "All" else client_data[client_data["period_year"].fillna(0).astype(int).astype(str) == selected_category]
    periods = periods_for_year(client_data, selected_category)
    if not periods:
        empty_state("Chưa có kỳ trong dữ liệu", icon="📅", subtitle="Hãy thử chọn công ty hoặc năm khác.")
        return

    pending_period = st.session_state.pop("pending_period_browser_period", None)
    if pending_period in periods:
        st.session_state.period_browser_selected_period = pending_period
    if "period_browser_selected_period" not in st.session_state or st.session_state.period_browser_selected_period not in periods:
        st.session_state.period_browser_selected_period = periods[0]

    selected_period = filter_cols[2].selectbox("Kỳ", periods, key="period_browser_selected_period")

    selected_data = year_data[year_data["period_label"].fillna("") == selected_period]
    st.subheader(f"{selected_client} - {selected_period}")
    st.markdown(
        f'<div class="summary-strip">{len(selected_data)} dòng · {selected_data["drawing_qty"].fillna(0).sum():g} drawings · SGD {selected_data["amount"].fillna(0).sum():,.0f}</div>',
        unsafe_allow_html=True,
    )

    panel_key = f"period_manage_{safe_filename(selected_client)}_{safe_filename(selected_period)}"
    edit_mode_key = f"{panel_key}_edit_mode"
    edit_mode = bool(st.session_state.get(edit_mode_key, False))
    action_cols = st.columns([1, 1, 1, 3])
    if action_cols[0].button("Hủy chỉnh sửa" if edit_mode else "Chỉnh sửa", width="stretch"):
        st.session_state[edit_mode_key] = not edit_mode
        st.rerun()
    if action_cols[1].button("Xuất PDF", type="primary", width="stretch", disabled=edit_mode):
        output_path = save_pdf_file(selected_data, f"{selected_client} - {selected_period}")
        st.toast(f"📄 Đã tạo PDF: {output_path.name}", icon="✅")
    if action_cols[2].button("Xuất Excel", width="stretch", disabled=edit_mode):
        output_path = save_excel_file(selected_data, f"{selected_client} - {selected_period}")
        st.toast(f"📊 Đã tạo Excel: {output_path.name}", icon="✅")

    if edit_mode:
        display_period_inline_edit_table(selected_data, panel_key)
    else:
        render_period_table(selected_data)


def display_period_drilldown_from_search(data: pd.DataFrame) -> None:
    if data.empty:
        return

    period_groups = []
    for (client_name, period_label), group in data.groupby(["client_name", "period_label"], dropna=False):
        if not period_label:
            continue
        period_groups.append((str(client_name or "GENERAL"), str(period_label), group))
    period_groups.sort(key=lambda item: (item[0], period_sort_key(item[1])), reverse=True)
    if not period_groups:
        return

    st.subheader("Mở kỳ từ kết quả tìm kiếm")
    columns = st.columns(3)
    for index, (client_label, period, period_data_for_button) in enumerate(period_groups):
        if columns[index % 3].button(
            f"{client_label} - {period}\n{len(period_data_for_button)} dự án",
            key=f"search_period_button_{client_label}_{period}",
            width="stretch",
        ):
            st.session_state.search_open_period = period
            st.session_state.search_open_client = client_label

    selected_period = st.session_state.get("search_open_period")
    if not selected_period:
        return
    selected_client = st.session_state.get("search_open_client", "GENERAL")
    period_data = data[
        (data["period_label"].fillna("") == selected_period)
        & (data["client_name"].fillna("GENERAL").map(uppercase_label) == selected_client)
    ]

    if st.session_state.get("search_open_period") == selected_period:
        st.caption(f"{selected_client} | {parse_year(selected_period) or ''} | {selected_period}")
        display_metrics(period_data)
        render_period_table(period_data)


def display_edit_delete_panel(data: pd.DataFrame, title: str = "Chọn dòng cần sửa hoặc xóa", selected_entry_id: int | None = None) -> None:
    if data.empty:
        st.info("Không có dòng dữ liệu để sửa/xóa.")
        return
    if st.session_state.pop("entry_saved_message", False):
        st.toast("💾 Đã lưu thay đổi.", icon="✅")

    display_data = data.sort_values(["period_label", "source_file", "source_row", "id"], na_position="last")
    if selected_entry_id is None:
        options = {}
        for period_label, period_rows in display_data.groupby("period_label", sort=False):
            period_rows = period_rows.sort_values(["source_file", "source_row", "id"], na_position="last").reset_index(drop=True)
            for index, row in period_rows.iterrows():
                row_no = index + 1
                options[f"#{int(row.id)} | No {row_no} | {period_label} | {uppercase_project_name(row.project_name)} | SGD {number_or_zero(row.amount):,.0f}"] = int(row.id)
        selected_label = st.selectbox(title, list(options.keys()))
        selected_entry_id = options[selected_label]
    else:
        st.markdown(f"**{title}**")

    entry = get_entry(int(selected_entry_id))
    if not entry:
        st.warning("Dòng này không còn tồn tại.")
        return

    config = get_company_owner_config(data)
    companies = [uppercase_label(value) for value in list(config.get("companies") or [])]
    owners_by_company = {uppercase_label(company): [uppercase_label(owner) for owner in owners] for company, owners in dict(config.get("owners_by_company") or {}).items()}
    entry_client = uppercase_label(entry.get("client_name")) or "GENERAL"
    if entry_client not in companies:
        companies = [entry_client] + companies
    owner_options = list(owners_by_company.get(entry_client) or [entry_client])
    entry_owner = uppercase_label(entry.get("owner"))
    if entry_owner and entry_owner not in owner_options:
        owner_options.append(entry_owner)

    with st.form(f"edit_entry_{entry['id']}"):
        top_cols = st.columns([1.2, 1.6, 1.2])
        client_name = top_cols[0].selectbox("Công ty", companies, index=companies.index(entry_client))
        project_name = top_cols[1].text_input("Dự án", value=uppercase_project_name(entry.get("project_name")))
        period_label = top_cols[2].text_input("Kỳ", value=str(entry.get("period_label") or ""))

        row_cols = st.columns([1, 2.2, 0.8, 0.9, 1.8])
        owner = row_cols[0].selectbox(
            "Người phụ trách",
            owner_options,
            index=owner_options.index(entry_owner) if entry_owner in owner_options else 0,
        )
        description = row_cols[1].text_area("Mô tả", value=str(entry.get("description") or ""), height=90)
        drawing_qty = row_cols[2].number_input("Số lượng", min_value=0.0, value=number_or_zero(entry.get("drawing_qty")), step=1.0)
        unit_price = row_cols[3].number_input("Đơn giá", min_value=0.0, value=number_or_zero(entry.get("unit_price")), step=10.0)
        amount = drawing_qty * unit_price
        current_image_path = str(entry.get("image_path") or "")
        with row_cols[4]:
            uploaded_files = st.file_uploader(
                "Hình ảnh",
                type=["png", "jpg", "jpeg", "webp"],
                accept_multiple_files=True,
                key=f"edit_images_{entry['id']}",
            )
        uploaded_paths = save_uploaded_images(uploaded_files) if uploaded_files else []
        existing_paths = image_paths_from_value(current_image_path)
        all_image_paths = existing_paths + [path for path in uploaded_paths if path not in existing_paths]
        image_path = "|".join(all_image_paths)
        if all_image_paths:
            uploaded_images_preview(all_image_paths, max_images=9)
        st.caption(f"Tổng tiền tự tính: SGD {amount:,.0f}")
        notes = st.text_area("Ghi chú", value=str(entry.get("notes") or ""))
        save_col, delete_col = st.columns(2)
        save_clicked = save_col.form_submit_button("Lưu thay đổi", type="primary", width="stretch")
        delete_clicked = delete_col.form_submit_button("Xóa dòng này", width="stretch")

        if save_clicked:
            if not project_name.strip():
                st.error("Tên dự án không được để trống.")
            else:
                update_entry(
                    int(entry["id"]),
                    {
                        "client_name": uppercase_label(client_name) or "GENERAL",
                        "project_name": project_name.strip(),
                        "owner": uppercase_label(owner),
                        "period_label": period_label.strip(),
                        "description": description.strip(),
                        "drawing_qty": drawing_qty,
                        "unit_price": unit_price,
                        "amount": amount,
                        "image_path": image_path.strip(),
                        "notes": notes.strip(),
                    },
                )
                st.session_state.entry_saved_message = True
                st.rerun()

        if delete_clicked:
            soft_delete_entry(int(entry["id"]))
            st.toast("🗑️ Đã xóa dòng khỏi app.", icon="✅")
            st.rerun()


def display_period_management_panel(data: pd.DataFrame, panel_key: str) -> None:
    if data.empty:
        return

    st.markdown("**Quản lý dòng trong kỳ**")
    sorted_data = data.sort_values(["source_file", "source_row", "id"], na_position="last").reset_index(drop=True)
    management_rows = pd.DataFrame(
        [
            {
                "Select": False,
                "No": index + 1,
                "Dự án": uppercase_project_name(row.get("project_name")),
                "Người phụ trách": uppercase_label(row.get("owner")),
                "Số lượng": number_or_zero(row.get("drawing_qty")),
                "Tổng tiền": number_or_zero(row.get("amount")),
                "id": int(row.get("id")),
            }
            for index, row in sorted_data.iterrows()
        ]
    )

    edited = st.data_editor(
        management_rows,
        key=f"{panel_key}_bulk_editor",
        width="stretch",
        hide_index=True,
        disabled=["No", "Dự án", "Người phụ trách", "Số lượng", "Tổng tiền", "id"],
        column_config={
            "Select": st.column_config.CheckboxColumn("Chọn", width="small"),
            "No": st.column_config.NumberColumn("No", width="small"),
            "Dự án": st.column_config.TextColumn("Dự án", width="large"),
            "Người phụ trách": st.column_config.TextColumn("Người phụ trách", width="small"),
            "Số lượng": st.column_config.NumberColumn("Số lượng", width="small"),
            "Tổng tiền": st.column_config.NumberColumn("Tổng tiền", format="SGD %.0f", width="medium"),
            "id": None,
        },
    )

    selected_ids = edited.loc[edited["Select"], "id"].astype(int).tolist() if not edited.empty else []
    action_cols = st.columns([1.1, 1.1, 3])
    action_cols[0].markdown(f"Đã chọn **{len(selected_ids)}** dòng")
    if action_cols[1].button("Xóa dòng đã chọn", disabled=not selected_ids, type="primary", width="stretch"):
        st.session_state[f"{panel_key}_confirm_delete_ids"] = selected_ids

    confirm_ids = st.session_state.get(f"{panel_key}_confirm_delete_ids", [])
    if confirm_ids:
        st.warning(f"Xác nhận xóa {len(confirm_ids)} dòng đã chọn khỏi app?")
        confirm_cols = st.columns([1, 1, 4])
        if confirm_cols[0].button("Xác nhận xóa", key=f"{panel_key}_confirm_delete", type="primary", width="stretch"):
            deleted_count = soft_delete_entries(confirm_ids)
            st.session_state.pop(f"{panel_key}_confirm_delete_ids", None)
            st.toast(f"🗑️ Đã xóa {deleted_count} dòng.", icon="✅")
            st.rerun()
        if confirm_cols[1].button("Hủy", key=f"{panel_key}_cancel_delete", width="stretch"):
            st.session_state.pop(f"{panel_key}_confirm_delete_ids", None)
            st.rerun()

    edit_options = {
        f"No {int(row['No'])} | {row['Dự án']} | SGD {number_or_zero(row['Tổng tiền']):,.0f}": int(row["id"])
        for _, row in management_rows.iterrows()
    }
    selected_label = st.selectbox("Chọn một dòng để sửa", [""] + list(edit_options.keys()), key=f"{panel_key}_edit_select")
    if selected_label:
        display_edit_delete_panel(data, title="Sửa dòng đã chọn", selected_entry_id=edit_options[selected_label])


def display_period_inline_edit_table(data: pd.DataFrame, panel_key: str) -> None:
    if data.empty:
        st.info("Không có dòng dữ liệu để chỉnh sửa.")
        return

    with st.container(border=True):
        _display_period_inline_edit_table_body(data, panel_key)


def _display_period_inline_edit_table_body(data: pd.DataFrame, panel_key: str) -> None:

    config = get_company_owner_config(data)
    owners_by_company = dict(config.get("owners_by_company") or {})
    client_name = uppercase_label(data.iloc[0].get("client_name")) or "GENERAL"
    owner_options = list(owners_by_company.get(client_name) or [client_name])
    for owner in unique_clean([uppercase_label(value) for value in data["owner"].dropna().tolist()]):
        if owner not in owner_options:
            owner_options.append(owner)

    sorted_data = data.sort_values(["source_file", "source_row", "id"], na_position="last").reset_index(drop=True)
    period_label = str(data.iloc[0].get("period_label") or "")
    st.markdown("**Chỉnh sửa trực tiếp**")
    st.markdown(
        f'<div class="summary-strip">Đang chỉnh sửa: {escape(client_name)} - {escape(period_label)} · {len(data)} dòng</div>',
        unsafe_allow_html=True,
    )
    header_cols = st.columns([2.2, 1, 2.4, 0.75, 0.9, 1.8, 0.35])
    for col, label in zip(header_cols, ["Dự án", "Người phụ trách", "Mô tả", "Số lượng", "Đơn giá", "Hình ảnh", ""]):
        col.caption(label)

    delete_marks_key = f"{panel_key}_delete_marks"
    if delete_marks_key not in st.session_state:
        st.session_state[delete_marks_key] = {}
    delete_marks = st.session_state[delete_marks_key]

    edited_entries: list[dict[str, object]] = []
    for _, row in sorted_data.iterrows():
        entry_id = int(row.get("id"))
        row_key = f"{panel_key}_{entry_id}"
        is_marked_delete = bool(delete_marks.get(str(entry_id), False))
        cols = st.columns([2.2, 1, 2.4, 0.75, 0.9, 1.8, 0.35])

        project_name = cols[0].text_input(
            "Dự án",
            value=uppercase_project_name(row.get("project_name")),
            key=f"{row_key}_project",
            label_visibility="collapsed",
            disabled=is_marked_delete,
        )
        current_owner = uppercase_label(row.get("owner")) or owner_options[0]
        current_owner_options = owner_options if current_owner in owner_options else owner_options + [current_owner]
        owner = cols[1].selectbox(
            "Người phụ trách",
            current_owner_options,
            index=current_owner_options.index(current_owner),
            key=f"{row_key}_owner",
            label_visibility="collapsed",
            disabled=is_marked_delete,
        )
        description = cols[2].text_area(
            "Mô tả",
            value=str(row.get("description") or ""),
            key=f"{row_key}_desc",
            label_visibility="collapsed",
            height=72,
            disabled=is_marked_delete,
        )
        drawing_qty = cols[3].number_input(
            "Số lượng",
            min_value=0.0,
            value=number_or_zero(row.get("drawing_qty")),
            step=1.0,
            key=f"{row_key}_qty",
            label_visibility="collapsed",
            disabled=is_marked_delete,
        )
        unit_price = cols[4].number_input(
            "Đơn giá",
            min_value=0.0,
            value=number_or_zero(row.get("unit_price")),
            step=10.0,
            key=f"{row_key}_unit",
            label_visibility="collapsed",
            disabled=is_marked_delete,
        )
        current_paths = image_paths_from_value(row.get("image_path"))
        with cols[5]:
            uploaded_files = st.file_uploader(
                "Hình ảnh",
                type=["png", "jpg", "jpeg", "webp"],
                accept_multiple_files=True,
                key=f"{row_key}_uploads",
                label_visibility="collapsed",
                disabled=is_marked_delete,
            )
            uploaded_paths = save_uploaded_images(uploaded_files) if uploaded_files and not is_marked_delete else []
            all_paths = current_paths + [path for path in uploaded_paths if path not in current_paths]
            if all_paths:
                uploaded_images_preview(all_paths, max_images=6)

        if cols[6].button("↩" if is_marked_delete else "X", key=f"{row_key}_delete", help="Bỏ đánh dấu xóa" if is_marked_delete else "Đánh dấu xóa", width="stretch"):
            delete_marks[str(entry_id)] = not is_marked_delete
            st.session_state[delete_marks_key] = delete_marks
            st.rerun()

        if is_marked_delete:
            st.warning(f"Dòng '{uppercase_project_name(row.get('project_name')) or entry_id}' sẽ bị xóa khi bấm Lưu tất cả thay đổi.")

        edited_entries.append(
            {
                "id": entry_id,
                "delete": is_marked_delete,
                "values": {
                    "client_name": uppercase_label(row.get("client_name")) or "GENERAL",
                    "project_name": project_name.strip(),
                    "owner": uppercase_label(owner),
                    "period_label": row.get("period_label") or "",
                    "description": description.strip(),
                    "drawing_qty": drawing_qty,
                    "unit_price": unit_price,
                    "amount": drawing_qty * unit_price,
                    "image_path": "|".join(all_paths),
                    "notes": row.get("notes") or "",
                },
            }
        )

    action_cols = st.columns([1.2, 1, 4])
    if action_cols[0].button("Lưu tất cả thay đổi", type="primary", width="stretch"):
        delete_ids = [entry["id"] for entry in edited_entries if entry["delete"]]
        update_count = 0
        for entry in edited_entries:
            if entry["delete"]:
                continue
            values = entry["values"]
            if not str(values["project_name"]).strip():
                st.error("Tên dự án không được để trống.")
                return
            update_entry(int(entry["id"]), values)
            update_count += 1
        deleted_count = soft_delete_entries(delete_ids)
        st.session_state.pop(delete_marks_key, None)
        st.session_state[f"{panel_key}_edit_mode"] = False
        st.toast(f"✅ Đã lưu {update_count} dòng và xóa {deleted_count} dòng.", icon="🚀")
        st.rerun()

    if action_cols[1].button("Hủy", width="stretch"):
        st.session_state.pop(delete_marks_key, None)
        st.session_state[f"{panel_key}_edit_mode"] = False
        st.rerun()


def display_search_table(data: pd.DataFrame, key: str = "search_table") -> int | None:
    if data.empty:
        st.info("Không tìm thấy dữ liệu phù hợp.")
        st.session_state.pop(f"{key}_selected_id", None)
        return None

    view = data[["id", "client_name", "period_label", "project_name", "owner", "description", "drawing_qty", "unit_price", "amount"]].copy().reset_index(drop=True)
    view["project_name"] = view["project_name"].map(uppercase_project_name)
    view["client_name"] = view["client_name"].map(uppercase_label)
    view["owner"] = view["owner"].map(uppercase_label)
    event = st.dataframe(
        view.drop(columns=["id"]),
        width="stretch",
        hide_index=True,
        on_select="rerun",
        selection_mode="single-row",
        key=key,
        column_config={
            "client_name": st.column_config.TextColumn("Công ty", width="small"),
            "period_label": st.column_config.TextColumn("Kỳ", width="medium"),
            "project_name": st.column_config.TextColumn("Dự án", width="large"),
            "description": st.column_config.TextColumn("Mô tả", width="large"),
            "amount": st.column_config.NumberColumn("Tổng tiền", format="SGD %.0f"),
        },
    )
    selected_rows = event.selection.rows if event and event.selection else []
    if selected_rows:
        selected_id = int(view.iloc[selected_rows[0]]["id"])
        selected_period = str(view.iloc[selected_rows[0]]["period_label"])
        selected_client = uppercase_label(view.iloc[selected_rows[0]]["client_name"]) or "GENERAL"
        st.session_state.search_open_period = selected_period
        st.session_state.search_open_client = selected_client
        st.session_state[f"{key}_selected_id"] = selected_id
        return selected_id
    st.session_state.pop(f"{key}_selected_id", None)
    return None


def search_export_title(data: pd.DataFrame, search_keyword: str, client_filter: str) -> str:
    keyword = normalize_text(search_keyword)
    if client_filter != "All":
        prefix = client_filter
    else:
        companies = unique_clean([uppercase_label(value) for value in data["client_name"].dropna().tolist()]) if not data.empty else []
        prefix = companies[0] if len(companies) == 1 else "Search"
    if keyword:
        return f'{prefix} - Search Summary - "{keyword}"'
    return f"{prefix} - Search Results"


def display_grouped_search_results(data: pd.DataFrame, search_keyword: str = "", client_filter: str = "All") -> None:
    if data.empty:
        empty_state("Không tìm thấy dữ liệu phù hợp", icon="🔍", subtitle="Hãy thử thay đổi từ khóa hoặc bộ lọc.")
        return

    sorted_data = data.sort_values(["client_name", "period_year", "period_label", "source_file", "source_row", "id"], ascending=[True, False, False, True, True, True])
    option_by_label: dict[str, int] = {}
    for _, row in sorted_data.iterrows():
        row_id = int(row["id"])
        label = (
            f"{uppercase_label(row.get('client_name')) or 'GENERAL'} · {row.get('period_label') or '-'} · "
            f"{uppercase_project_name(row.get('project_name')) or '-'} · SGD {number_or_zero(row.get('amount')):,.0f} · #{row_id}"
        )
        option_by_label[label] = row_id

    selection_key = "search_export_multiselect"
    valid_options = list(option_by_label.keys())
    current_selection = [value for value in st.session_state.get(selection_key, []) if value in option_by_label]
    if current_selection != st.session_state.get(selection_key, []):
        st.session_state[selection_key] = current_selection

    action_cols = st.columns([1, 1, 1, 3])
    if action_cols[0].button("Chọn tất cả kết quả", key="search_select_all_results", width="stretch"):
        st.session_state[selection_key] = valid_options
        st.rerun()
    if action_cols[1].button("Bỏ chọn", key="search_clear_export_selection", width="stretch"):
        st.session_state[selection_key] = []
        st.rerun()

    selected_labels = st.multiselect(
        "Chọn dòng để xuất",
        valid_options,
        default=st.session_state.get(selection_key, []),
        key=selection_key,
        placeholder="Chọn các dòng cần xuất PDF / Excel",
    )
    selected_ids = {option_by_label[label] for label in selected_labels if label in option_by_label}
    selected_data = data[data["id"].isin(selected_ids)] if selected_ids else data.iloc[0:0]

    total_qty = selected_data["drawing_qty"].fillna(0).sum() if not selected_data.empty else 0
    total_amount = selected_data["amount"].fillna(0).sum() if not selected_data.empty else 0
    export_title = search_export_title(selected_data, search_keyword, client_filter)
    bar_cols = st.columns([2.5, 1, 1])
    bar_cols[0].markdown(f"**Đã chọn:** {len(selected_data)} dòng · {total_qty:g} drawings · SGD {total_amount:,.0f}")
    if bar_cols[1].button("Xuất Excel", key="search_export_excel", width="stretch", disabled=selected_data.empty):
        output_path = save_excel_file(selected_data, export_title, include_period=True)
        st.toast(f"📊 Đã tạo Excel: {output_path.name}", icon="✅")
    if bar_cols[2].button("Xuất PDF", key="search_export_pdf", type="primary", width="stretch", disabled=selected_data.empty):
        output_path = save_pdf_file(selected_data, export_title, include_period=True, include_images=False)
        st.toast(f"📄 Đã tạo PDF: {output_path.name}", icon="✅")

    if not selected_data.empty:
        with st.container(border=True):
            st.markdown("**Search Summary Preview**")
            render_period_table(selected_data)

    grouped_data = sorted_data
    for client_name, client_rows in grouped_data.groupby("client_name", dropna=False):
        client_label = uppercase_label(client_name) or "GENERAL"
        with st.container(border=True):
            st.markdown(f"**{client_label}**")
            for period_label, period_rows in client_rows.groupby("period_label", dropna=False, sort=False):
                period_text = str(period_label or "")
                period_amount = period_rows["amount"].fillna(0).sum()
                period_qty = period_rows["drawing_qty"].fillna(0).sum()
                st.markdown(f"**{period_text}** · {len(period_rows)} dòng · {period_qty:g} drawings · SGD {period_amount:,.0f}")

                for _, row in period_rows.iterrows():
                    row_cols = st.columns([0.58, 0.16, 0.12, 0.14])
                    owner = uppercase_label(row.get("owner"))
                    owner_text = f" · {owner}" if owner else ""
                    selected_marker = " ✓" if int(row["id"]) in selected_ids else ""
                    row_cols[0].markdown(f"**{uppercase_project_name(row.get('project_name'))}**{owner_text}{selected_marker}")
                    row_cols[1].markdown(f"{number_or_zero(row.get('drawing_qty')):g} drawings")
                    row_cols[2].markdown(f"SGD {number_or_zero(row.get('unit_price')):,.0f}")
                    row_cols[3].markdown(f"**SGD {number_or_zero(row.get('amount')):,.0f}**")


def default_add_row() -> dict[str, object]:
    if "_add_row_counter" not in st.session_state:
        st.session_state._add_row_counter = 0
    st.session_state._add_row_counter += 1
    return {
        "_row_id": st.session_state._add_row_counter,
        "Project Name": "",
        "Owner": "ERNEST",
        "Description": "",
        "Drawings Quantity": 1.0,
        "Unit Price (SGD)": 110.0,
        "Reference Images": "",
        "Started In": "",
    }


def sync_add_rows_from_widgets() -> None:
    for row in st.session_state.get("add_rows", []):
        row_id = row.get("_row_id")
        if row_id is None:
            continue
        widget_map = {
            "Project Name": f"add_project_{row_id}",
            "Owner": f"add_owner_{row_id}",
            "Description": f"add_desc_{row_id}",
            "Drawings Quantity": f"add_qty_{row_id}",
            "Unit Price (SGD)": f"add_unit_{row_id}",
        }
        for field, key in widget_map.items():
            if key in st.session_state:
                row[field] = st.session_state[key]


def apply_project_history_to_add_row(row_id: int, project_name: str, history: dict[str, object], client_name: str) -> None:
    row_ref: dict[str, object] | None = None
    for row in st.session_state.get("add_rows", []):
        if row.get("_row_id") == row_id:
            row_ref = row
            break
    if row_ref is None:
        return

    project_key = f"add_project_{row_id}"
    owner_key = f"add_owner_{row_id}"
    desc_key = f"add_desc_{row_id}"
    unit_key = f"add_unit_{row_id}"

    latest_owner = uppercase_label(history.get("latest_owner") or row_ref.get("Owner") or client_name)
    latest_unit_price = number_or_zero(history.get("latest_unit_price"))
    current_desc = str(st.session_state.get(desc_key, row_ref.get("Description") or "") or "").strip()
    first_period = str(history.get("first_period") or "").strip()

    row_ref["Project Name"] = uppercase_project_name(project_name)
    row_ref["Owner"] = latest_owner
    row_ref["Description"] = current_desc
    row_ref["Unit Price (SGD)"] = latest_unit_price
    row_ref["Started In"] = first_period

    st.session_state[project_key] = row_ref["Project Name"]
    st.session_state[owner_key] = latest_owner
    st.session_state[desc_key] = current_desc
    st.session_state[unit_key] = latest_unit_price


def uploaded_images_preview(paths: list[str], max_images: int = 9) -> None:
    visible_paths = paths[:max_images]
    extra_count = max(0, len(paths) - max_images)
    cells = []
    for path_text in visible_paths:
        image_uri = image_data_uri(path_text)
        original_uri = original_image_data_uri(path_text)
        if image_uri:
            cells.append(
                f'<div class="upload-thumb image-open" data-full="{escape(original_uri or image_uri, quote=True)}">'
                f'<img src="{image_uri}" />'
                f'</div>'
            )
    if extra_count:
        cells.append(f'<div class="upload-thumb upload-more">+{extra_count}</div>')
    if not cells:
        st.caption("Chưa có ảnh upload.")
        return
    st.markdown(
        f'<div class="upload-grid">{"".join(cells)}</div>',
        unsafe_allow_html=True,
    )


def display_add_entry_panel(data: pd.DataFrame) -> None:
    saved_period = st.session_state.pop("add_saved_period", None)
    saved_client = st.session_state.pop("add_saved_client", None)
    saved_count = st.session_state.pop("add_saved_count", None)
    if saved_period and saved_client:
        saved_data = load_entries()
        saved_period_data = saved_data[
            (saved_data["client_name"].fillna("GENERAL").map(uppercase_label) == uppercase_label(saved_client))
            & (saved_data["period_label"].fillna("") == saved_period)
        ]
        st.toast(f"✅ Đã thêm {saved_count or len(saved_period_data)} dự án vào {saved_client} - {saved_period}.", icon="🚀")
        render_period_table(saved_period_data)

    config = get_company_owner_config(data)
    configured_companies = [uppercase_label(value) for value in list(config.get("companies") or ["GENERAL"])]
    owners_by_company = dict(config.get("owners_by_company") or {})
    current_client = configured_companies[0] if configured_companies else "GENERAL"
    if "add_rows" not in st.session_state:
        st.session_state.add_rows = [default_add_row()]
    sync_add_rows_from_widgets()

    top_cols = st.columns(3)
    with top_cols[0]:
        with st.container(border=True):
            st.markdown("**Công ty**")
            client_name = st.selectbox("Công ty", configured_companies, index=0, label_visibility="collapsed")
    client_name = uppercase_label(client_name) or "GENERAL"
    client_data = data[data["client_name"].fillna("GENERAL").map(uppercase_label) == client_name] if not data.empty else data
    categories = year_options(client_data)
    current_year = str(datetime.now().year)
    if current_year not in categories:
        categories = [current_year] + categories

    with top_cols[1]:
        with st.container(border=True):
            st.markdown("**Năm**")
            selected_category = st.selectbox("Năm", categories, key="add_category_year", label_visibility="collapsed")

    available_periods = periods_for_year(client_data, selected_category)
    with top_cols[2]:
        with st.container(border=True):
            st.markdown("**Kỳ**")
            mode = st.segmented_control("Chế độ", ["Chọn kỳ có sẵn", "Tạo kỳ mới"], default="Chọn kỳ có sẵn", label_visibility="collapsed")
            if mode == "Chọn kỳ có sẵn" and available_periods:
                period_label = st.selectbox("Kỳ", available_periods, key="add_existing_period", label_visibility="collapsed")
            else:
                period_label = st.text_input(
                    "Kỳ mới",
                    value=f"{datetime.now().strftime('%B').upper()} - {selected_category}",
                    label_visibility="collapsed",
                )

    valid_rows_for_preview = [row for row in st.session_state.add_rows if str(row.get("Project Name") or "").strip()]
    qty_total = sum(number_or_zero(row.get("Drawings Quantity")) for row in valid_rows_for_preview)
    amount_total = sum(number_or_zero(row.get("Drawings Quantity")) * number_or_zero(row.get("Unit Price (SGD)")) for row in valid_rows_for_preview)
    st.markdown(
        f'<div class="summary-strip">{escape(uppercase_label(client_name) or "GENERAL")} · {escape(str(period_label).strip() or "-")} · {len(valid_rows_for_preview)} dự án · {qty_total:g} drawings · SGD {amount_total:,.0f}</div>',
        unsafe_allow_html=True,
    )

    delete_index: int | None = None
    with st.container(border=True):
        st.markdown("**Dự án**")
        st.markdown(
            f'<div class="summary-strip">Đang nhập: {escape(uppercase_label(client_name) or "GENERAL")} - {escape(str(period_label).strip() or "-")} · {len(st.session_state.add_rows)} dòng</div>',
            unsafe_allow_html=True,
        )
        header_cols = st.columns([2.2, 1, 2.4, 0.75, 0.9, 1.8, 0.35])
        for col, label in zip(header_cols, ["Dự án", "Người phụ trách", "Mô tả", "Số lượng", "Đơn giá", "Hình ảnh", ""]):
            col.caption(label)

        owner_options = list(owners_by_company.get(client_name) or [client_name])
        project_options = project_history_options(data, client_name)
        for index, row in enumerate(st.session_state.add_rows):
            row_id = row.get("_row_id", index)
            cols = st.columns([2.2, 1, 2.4, 0.75, 0.9, 1.8, 0.35])
            with cols[0]:
                row["Project Name"] = st.text_input(
                    "Dự án",
                    value=str(row.get("Project Name") or ""),
                    key=f"add_project_{row_id}",
                    label_visibility="collapsed",
                )
                project_query = str(row["Project Name"] or "").strip().lower()
                matched_projects = [
                    project
                    for project in project_options
                    if project_query and project_query in str(project).lower()
                ][:5]
                if matched_projects:
                    st.caption("Dự án cũ phù hợp")
                    for project in matched_projects:
                        history = project_history_summary(data, client_name, project)
                        if not history:
                            continue
                        history_rows = history.get("rows") or []
                        button_label = f"{project} · bắt đầu {history.get('first_period') or '-'} · {len(history_rows)} dòng"
                        st.button(
                            button_label,
                            key=f"apply_project_history_{row_id}_{safe_filename(project)}",
                            width="stretch",
                            on_click=apply_project_history_to_add_row,
                            args=(row_id, project, history, client_name),
                        )
                if row["Project Name"] and st.button("Xem lịch sử", key=f"add_history_btn_{row_id}", width="stretch"):
                    st.session_state[f"show_project_history_{row_id}"] = not st.session_state.get(f"show_project_history_{row_id}", False)
            current_owner = uppercase_label(row.get("Owner")) or owner_options[0]
            if current_owner not in owner_options:
                owner_options = owner_options + [current_owner]
            row["Owner"] = cols[1].selectbox(
                "Người phụ trách",
                owner_options,
                index=owner_options.index(current_owner),
                key=f"add_owner_{row_id}",
                label_visibility="collapsed",
            )
            started_in = str(row.get("Started In") or "").strip()
            if started_in and row["Project Name"]:
                cols[0].caption(f"{row['Project Name']} · {uppercase_label(row['Owner'])} (started in {started_in})")
            row["Description"] = cols[2].text_area("Mô tả", value=str(row.get("Description") or ""), key=f"add_desc_{row_id}", label_visibility="collapsed", height=68)
            row["Drawings Quantity"] = cols[3].number_input("Số lượng", min_value=0.0, value=number_or_zero(row.get("Drawings Quantity")), step=1.0, key=f"add_qty_{row_id}", label_visibility="collapsed")
            row["Unit Price (SGD)"] = cols[4].number_input("Đơn giá", min_value=0.0, value=number_or_zero(row.get("Unit Price (SGD)")), step=10.0, key=f"add_unit_{row_id}", label_visibility="collapsed")
            with cols[5]:
                uploaded_files = st.file_uploader(
                    "Kéo thả hoặc chọn ảnh (PNG, JPG, WebP)",
                    type=["png", "jpg", "jpeg", "webp"],
                    accept_multiple_files=True,
                    key=f"add_uploads_{row_id}",
                    label_visibility="visible",
                )
                uploaded_paths = save_uploaded_images(uploaded_files) if uploaded_files else []
                existing_paths = image_paths_from_value(row.get("Reference Images"))
                all_paths = existing_paths + [path for path in uploaded_paths if path not in existing_paths]
                row["Reference Images"] = "|".join(all_paths)
                if all_paths:
                    uploaded_images_preview(all_paths, max_images=6)
            if cols[6].button("X", key=f"delete_add_row_{row_id}", help="Xóa dòng này", width="stretch"):
                delete_index = index
            if st.session_state.get(f"show_project_history_{row_id}") and row["Project Name"]:
                history = project_history_summary(data, client_name, row["Project Name"])
                with st.container(border=True):
                    if not history:
                        st.caption("Chưa có lịch sử cho dự án này trong công ty đang chọn.")
                    else:
                        history_rows = history.get("rows") or []
                        st.markdown(
                            f"**{row['Project Name']} - lịch sử** · bắt đầu **{history.get('first_period') or '-'}** · {len(history_rows)} dòng"
                        )
                        for history_row in history_rows:
                            desc_text = str(history_row.get("description") or "").strip().replace("\n", " · ")
                            if len(desc_text) > 120:
                                desc_text = f"{desc_text[:117]}..."
                            st.caption(
                                f"{history_row.get('period_label') or '-'} · "
                                f"{number_or_zero(history_row.get('drawing_qty')):g} drawings · "
                                f"SGD {number_or_zero(history_row.get('amount')):,.0f}"
                                f"{' · ' + desc_text if desc_text else ''}"
                            )

        if st.button("+ Thêm dòng", width="stretch"):
            st.session_state.add_rows.append(default_add_row())
            st.rerun()

    if delete_index is not None:
        st.session_state.add_rows.pop(delete_index)
        if not st.session_state.add_rows:
            st.session_state.add_rows.append(default_add_row())
        st.rerun()

    valid_rows = [row for row in st.session_state.add_rows if str(row.get("Project Name") or "").strip()]

    _left_spacer, bottom_right = st.columns([1.25, 1], gap="large")
    with bottom_right:
        with st.container(border=True):
            st.markdown("**Tổng kết trước khi lưu**")
            st.markdown(f"Công ty: **{uppercase_label(client_name) or 'GENERAL'}**")
            st.markdown(f"Kỳ: **{str(period_label).strip() or '-'}**")
            st.markdown(f"Dự án hợp lệ: **{len(valid_rows)}**")
            st.markdown(f"Drawings Quantity: **{qty_total:g}**")
            st.markdown(f"Tổng tiền: **SGD {amount_total:,.0f}**")

            submit_clicked = st.button("Thêm tất cả dự án vào kỳ", type="primary", width="stretch")

    if submit_clicked:
        if not str(period_label).strip():
            st.error("Cần nhập kỳ.")
            return
        if not valid_rows:
            st.error("Cần nhập ít nhất 1 section / dự án.")
            return

        entries = []
        for row in valid_rows:
            drawing_qty = number_or_zero(row.get("Drawings Quantity"))
            unit_price = number_or_zero(row.get("Unit Price (SGD)"))
            amount = drawing_qty * unit_price
            description = str(row.get("Description") or "").strip()
            started_in = str(row.get("Started In") or "").strip()
            if started_in:
                started_line = f"started in {started_in}"
                if started_line.lower() not in description.lower():
                    description = f"{description}\n{started_line}".strip() if description else started_line
            entries.append(
                {
                    "project_name": str(row.get("Project Name") or "").strip(),
                    "client_name": uppercase_label(client_name) or "GENERAL",
                    "owner": uppercase_label(row.get("Owner")),
                    "period_label": str(period_label).strip(),
                    "description": description,
                    "drawing_qty": drawing_qty,
                    "unit_price": unit_price,
                    "amount": amount,
                    "image_path": str(row.get("Reference Images") or "").strip(),
                    "notes": "",
                }
            )

        inserted = insert_manual_entries(entries)
        st.session_state.add_rows = [default_add_row()]
        st.session_state.pending_period_browser_period = str(period_label).strip()
        st.session_state.add_saved_client = uppercase_label(client_name) or "GENERAL"
        st.session_state.add_saved_period = str(period_label).strip()
        st.session_state.add_saved_count = inserted
        st.rerun()


def display_trash_panel() -> None:
    st.markdown("##### 🗑 Thùng rác")
    deleted = load_deleted_entries()
    if deleted.empty:
        st.caption("Không có dòng nào đã xóa.")
        return

    total_count = len(deleted)
    page_size = 10
    if "trash_show_all" not in st.session_state:
        st.session_state.trash_show_all = False
    display_count = total_count if st.session_state.trash_show_all else min(page_size, total_count)
    display_data = deleted.head(display_count)

    head_cols = st.columns([3, 1.2, 1.2])
    with head_cols[0]:
        st.caption(
            f"📦 **{total_count}** dòng đã xóa"
            + (f" (hiển thị {display_count})" if total_count > page_size else "")
        )
    with head_cols[1]:
        if st.button("↩ Khôi phục top 10", key="trash_restore_all", width="stretch",
                     help="Khôi phục tất cả dòng đang hiển thị",
                     disabled=display_data.empty):
            ids = [int(r["id"]) for _, r in display_data.iterrows()]
            for entry_id in ids:
                restore_entry(entry_id)
            st.toast(f"✅ Đã khôi phục {len(ids)} dòng.", icon="↩")
            st.rerun()
    with head_cols[2]:
        if total_count > page_size:
            label = "▲ Thu gọn" if st.session_state.trash_show_all else "▼ Xem tất cả"
            if st.button(label, key="trash_toggle", width="stretch"):
                st.session_state.trash_show_all = not st.session_state.trash_show_all
                st.rerun()

    with st.container(height=min(60 + display_count * 72, 480), border=True):
        for _, row in display_data.iterrows():
            entry_id = int(row["id"])
            client = uppercase_label(row.get("client_name") or "GENERAL")
            period = str(row.get("period_label") or "-")
            project = uppercase_project_name(row.get("project_name") or "-")
            deleted_at = str(row.get("deleted_at") or "")[:10]
            amount = number_or_zero(row.get("amount"))
            short_project = project if len(project) <= 45 else project[:42] + "…"

            row_cols = st.columns([5, 1, 1])
            with row_cols[0]:
                st.markdown(
                    f"**{escape(short_project)}**  \n"
                    f"<span style='color:var(--muted);font-size:12px;'>"
                    f"{escape(client)} · {escape(period)} · SGD {amount:,.0f} · {escape(deleted_at)}"
                    f"</span>",
                    unsafe_allow_html=True,
                )
            with row_cols[1]:
                if st.button("↩", key=f"trash_restore_{entry_id}", width="stretch", help="Khôi phục"):
                    restore_entry(entry_id)
                    st.toast("✅ Đã khôi phục.", icon="↩")
                    st.rerun()
            with row_cols[2]:
                if st.session_state.get(f"trash_confirm_del_{entry_id}"):
                    if st.button("⚠️", key=f"trash_confirm_yes_{entry_id}", width="stretch",
                                 type="primary", help="Xác nhận xóa vĩnh viễn"):
                        permanently_delete_entry(entry_id)
                        st.session_state.pop(f"trash_confirm_del_{entry_id}", None)
                        st.toast("🗑 Đã xóa vĩnh viễn.", icon="⚠️")
                        st.rerun()
                else:
                    if st.button("🗑", key=f"trash_delete_{entry_id}", width="stretch", help="Xóa vĩnh viễn"):
                        st.session_state[f"trash_confirm_del_{entry_id}"] = True
                        st.rerun()
            if st.session_state.get(f"trash_confirm_del_{entry_id}"):
                st.warning(
                    f"⚠️ Sắp xóa vĩnh viễn: **{escape(short_project)}**",
                    icon="⚠️",
                )


def display_image_gallery_panel(data: pd.DataFrame) -> None:
    st.markdown("##### \U0001f5bc\ufe0f Thư viện hình ảnh")
    st.caption("Click vào **tên dự án** để mở kỳ thanh toán · Click vào **hình** để phóng to")

    if data.empty:
        empty_state("Chưa có dữ liệu", icon="\U0001f5bc\ufe0f", subtitle="Vào Cài đặt > Import dữ liệu để bắt đầu.")
        return

    image_rows = data[data["image_path"].fillna("").str.strip() != ""].copy()
    if image_rows.empty:
        empty_state("Chưa có hình ảnh nào", icon="\U0001f5bc\ufe0f", subtitle="Các dòng có hình ảnh sẽ hiển thị tại đây.")
        return

    # Filters
    filter_cols = st.columns([1.2, 1, 2])
    clients = ["All"] + client_options(image_rows)
    with filter_cols[0]:
        sel_client = st.selectbox("Công ty", clients, key="gallery_client")
    scoped = image_rows.copy()
    if sel_client != "All":
        scoped = scoped[scoped["client_name"].fillna("GENERAL").map(uppercase_label) == sel_client]
    years = ["All"] + year_options(scoped)
    with filter_cols[1]:
        sel_year = st.selectbox("Năm", years, key="gallery_year")
    if sel_year != "All":
        scoped = scoped[scoped["period_year"].fillna(0).astype(int).astype(str) == sel_year]
    with filter_cols[2]:
        kw = st.text_input("Tìm theo tên dự án", placeholder="Ví dụ: Jalan Peminpin", key="gallery_keyword")
    if kw.strip():
        scoped = scoped[scoped["project_name"].fillna("").str.lower().str.contains(kw.strip().lower(), regex=False)]

    if scoped.empty:
        empty_state("Không tìm thấy hình ảnh phù hợp", icon="\U0001f50d", subtitle="Hãy thử thay đổi bộ lọc hoặc từ khóa.")
        return

    # Sort: newest period first
    scoped = scoped.copy()
    scoped["_psort"] = scoped["period_label"].fillna("").map(period_sort_key)
    scoped = scoped.sort_values(
        ["_psort", "client_name", "source_file", "source_row", "id"],
        ascending=[False, True, True, True, True],
        na_position="last",
    ).drop(columns=["_psort"])

    # Build flat list of items
    items = []
    for _, row in scoped.iterrows():
        paths = image_paths_from_value(row.get("image_path"))
        for path_text in paths:
            thumb_uri = image_data_uri(path_text)
            if not thumb_uri:
                continue
            original_uri = original_image_data_uri(path_text)
            items.append((row, thumb_uri, original_uri or thumb_uri))

    if not items:
        empty_state("Không tìm thấy hình ảnh phù hợp", icon="\U0001f50d", subtitle="Hãy thử thay đổi bộ lọc hoặc từ khóa.")
        return

    # Pagination
    page_size = 20
    total_pages = max(1, (len(items) - 1) // page_size + 1)
    page_key = "gallery_page"
    if page_key not in st.session_state:
        st.session_state[page_key] = 0
    if st.session_state[page_key] >= total_pages:
        st.session_state[page_key] = max(0, total_pages - 1)

    info_cols = st.columns([4, 1, 1, 1])
    with info_cols[0]:
        st.caption(
            f"\U0001f5bc\ufe0f **{len(items)}** hình từ **{len(scoped)}** dự án"
            + (f" · trang {st.session_state[page_key]+1}/{total_pages}" if total_pages > 1 else "")
        )
    if total_pages > 1:
        with info_cols[1]:
            if st.button("\u25c4", key="gallery_prev", disabled=st.session_state[page_key] <= 0, width="stretch"):
                st.session_state[page_key] -= 1
                st.rerun()
        with info_cols[2]:
            if st.button("\u25ba", key="gallery_next", disabled=st.session_state[page_key] >= total_pages - 1, width="stretch"):
                st.session_state[page_key] += 1
                st.rerun()
        with info_cols[3]:
            st.markdown(
                f'<div style="text-align:center;font-size:12px;color:var(--muted);padding-top:6px;">' +
                f'{st.session_state[page_key] + 1}/{total_pages}</div>',
                unsafe_allow_html=True,
            )

    page = st.session_state[page_key]
    page_items = items[page * page_size : (page + 1) * page_size]

    num_cols = 4
    import streamlit.components.v1 as components

    def render_gallery_image_row(chunk: list[tuple[pd.Series, str, str]], row_start: int) -> None:
        cells: list[str] = []
        for ci, (row, thumb_uri, original_uri) in enumerate(chunk):
            client = uppercase_label(row.get("client_name")) or "GENERAL"
            period = str(row.get("period_label") or "-")
            tooltip = f"{escape(client)} · {escape(period)}"
            cells.append(
                f'<button type="button" class="gcell" data-full="{escape(original_uri, quote=True)}" title="{tooltip}">'
                f'<img src="{thumb_uri}" alt="" />'
                f'</button>'
            )
        for _ in range(num_cols - len(chunk)):
            cells.append('<div class="gcell ghost"></div>')

        row_html = (
            '<!DOCTYPE html><html><head><meta charset="utf-8"><style>'
            ':root{--bd:#E8E8E4;--sf:#F7F7F5;--ac:#B8760A;--sa:#F6E8D0;}'
            'body{margin:0;padding:4px;background:transparent;}'
            '.grid{display:grid;grid-template-columns:repeat(' + str(num_cols) + ',1fr);gap:10px;}'
            '.gcell{height:190px;border:1.5px solid var(--bd);border-radius:10px;overflow:hidden;'
            'background:var(--sf);cursor:zoom-in;display:flex;align-items:center;justify-content:center;'
            'transition:border-color .15s,box-shadow .15s;padding:0;}'
            '.gcell:hover{border-color:var(--ac);box-shadow:0 0 0 2px var(--sa);}'
            '.gcell img{width:100%;height:100%;object-fit:contain;display:block;pointer-events:none;}'
            '.ghost{visibility:hidden;}'
            '</style></head><body>'
            '<div class="grid">' + ''.join(cells) + '</div>'
            '<script>'
            'function closeOverlay(){'
            '  var doc=window.parent&&window.parent.document?window.parent.document:document;'
            '  var old=doc.getElementById("galleryZoomOverlay");'
            '  if(old)old.remove();'
            '}'
            'function openOverlay(src){'
            '  var doc=window.parent&&window.parent.document?window.parent.document:document;'
            '  closeOverlay();'
            '  var overlay=doc.createElement("div");overlay.id="galleryZoomOverlay";'
            '  overlay.style.cssText="position:fixed;inset:0;z-index:2147483647;background:rgba(0,0,0,.9);display:flex;align-items:center;justify-content:center;padding:24px;cursor:zoom-out;";'
            '  var img=doc.createElement("img");img.src=src;img.style.cssText="max-width:96vw;max-height:92vh;object-fit:contain;border-radius:8px;box-shadow:0 20px 80px rgba(0,0,0,.6);";'
            '  var btn=doc.createElement("button");btn.innerHTML="&#215;";btn.type="button";'
            '  btn.style.cssText="position:fixed;top:12px;right:16px;width:44px;height:44px;border:0;border-radius:50%;background:rgba(255,255,255,.18);color:white;font-size:30px;line-height:44px;cursor:pointer;";'
            '  btn.onclick=function(e){e.stopPropagation();closeOverlay();};'
            '  overlay.onclick=closeOverlay;img.onclick=function(e){e.stopPropagation();};'
            '  overlay.appendChild(img);overlay.appendChild(btn);doc.body.appendChild(overlay);'
            '}'
            'document.querySelectorAll(".gcell[data-full]").forEach(function(cell){'
            '  cell.addEventListener("click",function(){openOverlay(cell.getAttribute("data-full"));});'
            '});'
            'document.addEventListener("keydown",function(e){if(e.key==="Escape")closeOverlay();});'
            '</script>'
            '</body></html>'
        )
        components.html(row_html, height=210, scrolling=False)

    for row_start in range(0, len(page_items), num_cols):
        chunk = page_items[row_start : row_start + num_cols]
        render_gallery_image_row(chunk, row_start)
        btn_cols = st.columns(num_cols, gap="small")
        for ci, (row, _t, _o) in enumerate(chunk):
            project = uppercase_project_name(row.get("project_name")) or "-"
            client = uppercase_label(row.get("client_name")) or "GENERAL"
            period = str(row.get("period_label") or "-")
            amount = number_or_zero(row.get("amount"))
            short_project = project if len(project) <= 32 else project[:30] + "…"
            with btn_cols[ci]:
                btn_key = f"gal_jump_{page}_{row_start}_{ci}_{safe_filename(project)[:16]}"
                if st.button(
                    short_project,
                    key=btn_key,
                    help=f"{client} · {period} → mở Kỳ thanh toán",
                    width="stretch",
                ):
                    _jump_to_period_browser(client, period)
                st.markdown(
                    f'<div style="font-size:11px;color:var(--muted);text-align:center;'
                    f'line-height:1.4;margin-top:-4px;margin-bottom:10px;">'
                    f'{escape(period)}<br/>'
                    f'<span style="color:var(--accent);font-weight:600;">SGD {amount:,.0f}</span></div>',
                    unsafe_allow_html=True,
                )


def display_settings_panel(data: pd.DataFrame) -> None:
    st.markdown("### ⚙️ Cài đặt")
    tab_company, tab_import, tab_sync, tab_payment, tab_backup, tab_trash = st.tabs([
        "🏢 Công ty & Owners",
        "📥 Import dữ liệu",
        "☁️ Đồng bộ",
        "💰 Thanh toán",
        "💾 Backup",
        "🗑 Thùng rác",
    ])
    with tab_company:
        _settings_company(data)
    with tab_import:
        _settings_import()
    with tab_sync:
        _settings_sync()
    with tab_payment:
        _settings_payment()
    with tab_backup:
        _settings_backup()
    with tab_trash:
        display_trash_panel()


def _settings_company(data: pd.DataFrame) -> None:
    config = get_company_owner_config(data)
    companies = list(config.get("companies") or [])
    owners_by_company = dict(config.get("owners_by_company") or {})
    st.caption("Quản lý danh sách công ty và người phụ trách tương ứng.")
    col_left, col_right = st.columns([1, 1.1], gap="medium")
    with col_left:
        with st.container(border=True):
            st.markdown("##### 🏢 Danh sách công ty")
            companies_text = st.text_area(
                "Mỗi dòng = 1 công ty",
                value="\n".join(companies),
                height=280,
                key="settings_companies_text",
                label_visibility="collapsed",
            )
            st.caption(f"📊 Hiện có: **{len(companies)}** công ty")
    clean_companies = unique_clean([uppercase_label(value) for value in companies_text.splitlines()])
    if not clean_companies:
        clean_companies = ["GENERAL"]
    with col_right:
        with st.container(border=True):
            st.markdown("##### 👤 Người phụ trách")
            selected_company = st.selectbox(
                "Chọn công ty",
                clean_companies,
                key="settings_owner_company",
            )
            current_owners = owners_by_company.get(selected_company, [selected_company])
            owners_text = st.text_area(
                "Mỗi dòng = 1 người",
                value="\n".join([uppercase_label(value) for value in current_owners]),
                height=240,
                key=f"settings_owners_text_{safe_filename(selected_company)}",
                label_visibility="collapsed",
            )
            parsed_owners = unique_clean([uppercase_label(value) for value in owners_text.splitlines()])
            st.caption(f"👥 Hiện có: **{len(parsed_owners)}** người")
    st.markdown("")
    action_cols = st.columns([1, 1, 4])
    with action_cols[0]:
        if st.button("💾 Lưu tất cả", type="primary", width="stretch", key="save_company_settings"):
            updated_owners = {
                company: list(owners_by_company.get(company, [company]))
                for company in clean_companies
            }
            updated_owners[selected_company] = parsed_owners
            save_company_owner_config(clean_companies, updated_owners)
            st.toast("✅ Đã lưu cài đặt.", icon="⚙️")
            st.rerun()
    with action_cols[1]:
        if st.button("🔄 Reset", width="stretch", key="reset_company_settings"):
            set_setting("company_owner_config", "")
            st.toast("🔄 Đã reset.", icon="↩")
            st.rerun()


def _settings_import() -> None:
    st.caption("Chọn folder chứa file Excel tổng kết, sau đó import hoặc cập nhật dữ liệu vào app.")
    with st.container(border=True):
        st.markdown("##### 📥 Import dữ liệu từ Excel")
        current_import_folder = get_setting("import_folder", str(TONG_KET_DIR))
        import_folder = st.text_input(
            "Folder import Excel",
            value=current_import_folder,
            key="settings_import_folder",
        )
        folder_path = Path(import_folder)
        if folder_path.exists() and folder_path.is_dir():
            st.caption(f"Folder đang dùng: `{folder_path}`")
        else:
            st.warning("Folder import không tồn tại.")

        if st.button("Import / cập nhật từ Excel", type="primary", width="stretch", key="settings_import_excel"):
            if not folder_path.exists() or not folder_path.is_dir():
                st.error("Không thể import vì folder không tồn tại.")
            else:
                progress_bar = st.progress(0, text="Chuẩn bị import...")

                def update_progress(idx, total, file_path):
                    pct = int((idx / total) * 100)
                    progress_bar.progress(pct, text=f"Đang đọc ({idx}/{total}): {file_path.name}")

                with st.spinner("Đang import dữ liệu từ Excel..."):
                    set_setting("import_folder", str(folder_path))
                    count = import_excel_data(folder_path, progress_callback=update_progress)

                progress_bar.empty()
                clear_entries_cache()
                st.toast(f"✅ Đã import/cập nhật {count} dòng từ Excel.", icon="🚀")
                st.rerun()


def run_git_command(args: list[str]) -> tuple[bool, str]:
    try:
        result = subprocess.run(
            ["git", *args],
            cwd=APP_DIR,
            capture_output=True,
            text=True,
            timeout=120,
            check=False,
        )
    except Exception as exc:
        return False, str(exc)

    output = "\n".join(part.strip() for part in [result.stdout, result.stderr] if part and part.strip())
    return result.returncode == 0, output or "(không có output)"


def _settings_sync() -> None:
    st.caption("Đồng bộ dữ liệu local lên GitHub để app Cloud read-only cập nhật theo nhánh main.")
    tracked_targets = [
        "app.py",
        "view.py",
        "requirements.txt",
        ".gitignore",
        "tong_ket.db",
        "uploads",
    ]

    with st.container(border=True):
        st.markdown("##### ☁️ Đồng bộ lên Internet")
        status_ok, status_output = run_git_command(["status", "--short"])
        remote_ok, remote_output = run_git_command(["remote", "-v"])
        branch_ok, branch_output = run_git_command(["branch", "--show-current"])

        if branch_ok:
            st.caption(f"Nhánh hiện tại: `{branch_output.strip() or '-'}`")
        if remote_ok:
            st.code(remote_output, language="text")
        if status_ok:
            st.text_area("Thay đổi đang chờ đồng bộ", value=status_output, height=160, disabled=True)
        else:
            st.error(status_output)

        commit_message = st.text_input(
            "Nội dung commit",
            value=f"Sync data {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            key="sync_commit_message",
        )

        if st.button("Đồng bộ lên Internet", type="primary", width="stretch", key="sync_to_github"):
            if not commit_message.strip():
                st.error("Cần nhập nội dung commit.")
                return

            steps: list[tuple[str, bool, str]] = []
            for args in [
                ["add", *tracked_targets],
                ["commit", "-m", commit_message.strip()],
                ["push", "origin", "main"],
            ]:
                ok, output = run_git_command(args)
                steps.append(("git " + " ".join(args), ok, output))
                if not ok:
                    # Nothing to commit is acceptable: still push, because remote may lag.
                    if args[0] == "commit" and "nothing to commit" in output.lower():
                        continue
                    break

            for command, ok, output in steps:
                st.markdown(("✅" if ok else "⚠️") + f" `{command}`")
                st.code(output, language="text")

            if steps and steps[-1][0].startswith("git push") and steps[-1][1]:
                st.toast("☁️ Đã đồng bộ lên GitHub. Streamlit Cloud sẽ tự cập nhật.", icon="✅")
            else:
                st.warning("Đồng bộ chưa hoàn tất. Xem output phía trên để biết bước bị lỗi.")


def _settings_payment() -> None:
    st.caption("Thông tin hiển thị cuối các file PDF xuất ra.")
    with st.container(border=True):
        st.markdown("##### 💰 Thông tin thanh toán")
        payment_info = st.text_area(
            "Nội dung",
            value=get_setting("payment_info", DEFAULT_PAYMENT_INFO),
            height=280,
            key="payment_info_text",
            label_visibility="collapsed",
        )
        action_cols = st.columns([1, 1, 4])
        with action_cols[0]:
            if st.button("💾 Lưu", type="primary", width="stretch", key="save_payment"):
                set_setting("payment_info", payment_info.strip())
                st.toast("💰 Đã lưu.", icon="✅")
        with action_cols[1]:
            if st.button("🔄 Reset", width="stretch", key="reset_payment"):
                set_setting("payment_info", DEFAULT_PAYMENT_INFO)
                st.toast("🔄 Đã reset.", icon="↩")
                st.rerun()


def _settings_backup() -> None:
    st.caption("Database tự động backup mỗi ngày. Bạn có thể tạo backup thủ công hoặc khôi phục từ bản cũ.")
    backup_files = sorted(BACKUP_DIR.glob("tong_ket_*.db"), key=lambda p: p.stat().st_mtime, reverse=True)

    action_cols = st.columns([1, 1, 1, 3])
    with action_cols[0]:
        if st.button("💾 Backup ngay", type="primary", width="stretch", key="manual_backup"):
            import shutil
            ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
            bp = BACKUP_DIR / f"tong_ket_{ts}.db"
            shutil.copy2(DB_PATH, bp)
            st.toast(f"💾 Đã tạo: {bp.name}", icon="✅")
            st.rerun()
    with action_cols[1]:
        if backup_files and st.button("♻️ Khôi phục gần nhất", width="stretch", key="restore_latest"):
            st.session_state["confirm_restore_latest"] = True
            st.rerun()
    with action_cols[2]:
        if len(backup_files) > 5 and st.button("🗑 Dọn backup cũ", width="stretch", key="delete_old"):
            st.session_state["confirm_delete_old"] = True
            st.rerun()

    if st.session_state.get("confirm_restore_latest") and backup_files:
        latest = backup_files[0]
        st.warning(
            f"⚠️ **CẢNH BÁO:** Khôi phục từ `{latest.name}` sẽ **GHI ĐÈ** database hiện tại. "
            f"DB hiện tại sẽ được backup an toàn trước.",
            icon="⚠️",
        )
        conf_cols = st.columns([1, 1, 4])
        with conf_cols[0]:
            if st.button("⚠️ Xác nhận", type="primary", key="confirm_restore_yes"):
                import shutil
                safety_ts = datetime.now().strftime("%Y-%m-%d_%H%M%S_pre")
                safety_path = BACKUP_DIR / f"tong_ket_{safety_ts}.db"
                shutil.copy2(DB_PATH, safety_path)
                shutil.copy2(latest, DB_PATH)
                clear_entries_cache()
                st.session_state.pop("confirm_restore_latest", None)
                st.toast(f"♻️ Đã khôi phục. Safety: {safety_path.name}", icon="✅")
                st.rerun()
        with conf_cols[1]:
            if st.button("Hủy", key="confirm_restore_no"):
                st.session_state.pop("confirm_restore_latest", None)
                st.rerun()

    if st.session_state.get("confirm_delete_old"):
        old_count = len(backup_files) - 5
        st.warning(f"⚠️ Sẽ xóa **{old_count}** bản cũ, giữ 5 bản mới nhất.")
        conf_cols = st.columns([1, 1, 4])
        with conf_cols[0]:
            if st.button("⚠️ Xác nhận xóa", type="primary", key="confirm_delete_yes"):
                for old in backup_files[5:]:
                    old.unlink(missing_ok=True)
                st.session_state.pop("confirm_delete_old", None)
                st.toast(f"🗑 Đã xóa {old_count} bản cũ.", icon="✅")
                st.rerun()
        with conf_cols[1]:
            if st.button("Hủy", key="confirm_delete_no"):
                st.session_state.pop("confirm_delete_old", None)
                st.rerun()

    st.markdown("")
    with st.container(border=True):
        st.markdown("##### 📋 Danh sách bản backup")
        if not backup_files:
            st.caption("Chưa có bản backup nào.")
        else:
            total_size = sum(bf.stat().st_size for bf in backup_files) / (1024 * 1024)
            st.caption(f"📦 **{len(backup_files)}** bản · Tổng **{total_size:.1f} MB**")
            with st.container(height=min(60 + min(len(backup_files), 10) * 56, 400), border=False):
                for bf in backup_files[:20]:
                    size_kb = bf.stat().st_size // 1024
                    mtime = datetime.fromtimestamp(bf.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
                    r_cols = st.columns([3, 2, 1, 1])
                    with r_cols[0]:
                        st.markdown(f"📦 **`{bf.name}`**")
                    with r_cols[1]:
                        st.caption(f"{size_kb} KB · {mtime}")
                    with r_cols[2]:
                        if st.button("♻️", key=f"restore_{bf.name}", help="Khôi phục", width="stretch"):
                            st.session_state[f"confirm_restore_{bf.name}"] = True
                            st.rerun()
                    with r_cols[3]:
                        if st.button("🗑", key=f"del_{bf.name}", help="Xóa bản này", width="stretch"):
                            bf.unlink(missing_ok=True)
                            st.toast(f"🗑 Đã xóa {bf.name}", icon="✅")
                            st.rerun()
                    if st.session_state.get(f"confirm_restore_{bf.name}"):
                        st.warning(f"⚠️ Khôi phục từ `{bf.name}`?")
                        cc = st.columns([1, 1, 4])
                        if cc[0].button("⚠️ Xác nhận", key=f"yes_{bf.name}", type="primary"):
                            import shutil
                            safety_ts = datetime.now().strftime("%Y-%m-%d_%H%M%S_pre")
                            safety_path = BACKUP_DIR / f"tong_ket_{safety_ts}.db"
                            shutil.copy2(DB_PATH, safety_path)
                            shutil.copy2(bf, DB_PATH)
                            clear_entries_cache()
                            st.session_state.pop(f"confirm_restore_{bf.name}", None)
                            st.toast(f"♻️ Đã khôi phục", icon="✅")
                            st.rerun()
                        if cc[1].button("Hủy", key=f"no_{bf.name}"):
                            st.session_state.pop(f"confirm_restore_{bf.name}", None)
                            st.rerun()
                if len(backup_files) > 20:
                    st.caption(f"... và {len(backup_files) - 20} bản nữa")


def main() -> None:
    st.set_page_config(page_title="Tong Ket Manager", page_icon="📁", layout="wide")
    inject_app_css()
    # Chuẩn hóa image path cũ -> relative
    _migrated = migrate_absolute_image_paths()
    if _migrated:
        st.toast(f"🔧 Đã chuẩn hóa {_migrated} image path.", icon="🔧")
    # Daily auto-backup
    _backup_result = auto_backup_db()
    if _backup_result:
        st.toast(f"💾 Auto-backup: {_backup_result.name}", icon="✅")
    st.title("Tong Ket Manager")
    st.caption("Quản lý project, drawing, revision và xuất PDF từ dữ liệu tổng kết.")

    data = load_entries()

    # Inject JS to auto-click a tab after gallery navigation jump
    jump_tab = st.session_state.pop("_jump_to_tab_label", None)
    if jump_tab:
        escaped_label = jump_tab.replace("'", "\\'")
        import streamlit.components.v1 as components
        components.html(
            f"""<script>
            (function() {{
                var doc = window.parent && window.parent.document ? window.parent.document : document;
                function clickTab() {{
                    var tabs = doc.querySelectorAll('[data-baseweb="tab"]');
                    for (var i = 0; i < tabs.length; i++) {{
                        if (tabs[i].textContent.trim().indexOf('{escaped_label}') !== -1) {{
                            tabs[i].click();
                            return true;
                        }}
                    }}
                    return false;
                }}
                var attempts = 0;
                var iv = setInterval(function() {{
                    if (clickTab() || ++attempts > 20) clearInterval(iv);
                }}, 80);
            }})();
            </script>""",
            height=0,
            scrolling=False,
        )
        st.toast(f"✅ Đã chuyển tới: {jump_tab}", icon="📂")

    tab_dashboard, tab_add, tab_periods, tab_search, tab_gallery, tab_export, tab_settings = st.tabs(
        ["Tổng quan", "Thêm mới", "Kỳ thanh toán", "Tìm kiếm", "Hình ảnh", "Xuất PDF", "Cài đặt"]
    )

    with tab_dashboard:
        display_bento_dashboard(data)

    config = get_company_owner_config(data)
    configured_companies = [uppercase_label(value) for value in list(config.get("companies") or [])]
    configured_owners = []
    for owner_list in dict(config.get("owners_by_company") or {}).values():
        configured_owners.extend([uppercase_label(value) for value in owner_list])
    clients = ["All"] + unique_clean(configured_companies + (client_options(data) if not data.empty else []))
    owners = ["All"] + unique_clean(configured_owners + ([uppercase_label(value) for value in data["owner"].dropna().unique().tolist()] if not data.empty else []))
    years = ["All"] + sorted([str(int(x)) for x in data["period_year"].dropna().unique().tolist()], reverse=True) if not data.empty else ["All"]
    periods = ["All"] + sorted([x for x in data["period_label"].dropna().unique().tolist() if x], reverse=True) if not data.empty else ["All"]

    with tab_add:
        display_add_entry_panel(data)

    with tab_periods:
        display_period_browser(data)

    with tab_search:
        with st.container(border=True):
            col0, col1, col2, col3, col4, col5 = st.columns([1.4, 2, 1, 1, 1, 1])
            quick_query = col0.text_input("Tìm nhanh", placeholder="Gõ để hiện gợi ý", key="search_autocomplete_text")
            keyword = col1.text_input("Tìm project / mô tả", placeholder="Ví dụ: Jalan Peminpin, West Coast, Doreen")
            client = col2.selectbox("Công ty", clients)
            owner = col3.selectbox("Người phụ trách", owners)
            year = col4.selectbox("Năm", years)
            period = col5.selectbox("Kỳ", periods)
            suggestion_keyword = ""
            suggestion_options = search_autocomplete_options(data, quick_query)
            if suggestion_options:
                st.caption("Gợi ý")
                suggestion_cols = st.columns(min(4, len(suggestion_options)))
                for index, (label, value) in enumerate(suggestion_options):
                    if suggestion_cols[index % len(suggestion_cols)].button(label, key=f"search_suggestion_btn_{index}_{safe_filename(value)}"):
                        suggestion_keyword = value
                        st.session_state.search_selected_keyword = value
            elif quick_query:
                st.caption("Không có gợi ý phù hợp.")
        selected_keyword = st.session_state.get("search_selected_keyword", "")
        if quick_query and selected_keyword and quick_query.lower() in selected_keyword.lower():
            active_keyword = selected_keyword
        else:
            active_keyword = quick_query or keyword
        filtered = filter_entries(data, active_keyword, client, owner, year, period)
        st.markdown(
            f'<div class="summary-strip">{len(filtered)} dòng · {filtered["project_name"].map(uppercase_project_name).nunique() if not filtered.empty else 0} dự án · SGD {filtered["amount"].fillna(0).sum() if not filtered.empty else 0:,.0f}</div>',
            unsafe_allow_html=True,
        )
        display_grouped_search_results(filtered, active_keyword, client)

    with tab_gallery:
        display_image_gallery_panel(data)

    with tab_export:
        col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
        export_keyword = col1.text_input("Lọc nội dung xuất PDF", key="export_keyword")
        export_client = col2.selectbox("Công ty xuất", clients, key="export_client")
        export_owner = col3.selectbox("Người phụ trách xuất", owners, key="export_owner")
        export_year = col4.selectbox("Năm xuất", years, key="export_year")
        export_data = filter_entries(data, export_keyword, export_client, export_owner, export_year, "All")
        display_metrics(export_data)
        default_title = f"{export_client} - Project Summary" if export_client != "All" else "Project Summary"
        title = st.text_input("Tiêu đề PDF", value=default_title)
        export_view = export_data[["client_name", "period_label", "project_name", "owner", "description", "drawing_qty", "unit_price", "amount"]].copy() if not export_data.empty else export_data
        if not export_view.empty:
            export_view["project_name"] = export_view["project_name"].map(uppercase_project_name)
            export_view["client_name"] = export_view["client_name"].map(uppercase_label)
            export_view["owner"] = export_view["owner"].map(uppercase_label)
        st.dataframe(
            export_view,
            width="stretch",
            hide_index=True,
        )
        col_prev, col_save = st.columns([1, 1])
        if col_prev.button("👁 Preview PDF", width="stretch", disabled=export_data.empty):
            st.session_state["pdf_preview_bytes"] = make_pdf(export_data, title)
        if col_save.button("💾 Lưu PDF ra file", type="primary", width="stretch", disabled=export_data.empty):
            output_path = save_pdf_file(export_data, title)
            st.toast(f"📄 Đã tạo PDF: {output_path.name}", icon="✅")
            st.session_state.pop("pdf_preview_bytes", None)

        pdf_bytes = st.session_state.get("pdf_preview_bytes")
        if pdf_bytes:
            import base64
            b64 = base64.b64encode(pdf_bytes).decode()
            pdf_html = (
                f'<iframe src="data:application/pdf;base64,{b64}" ' +
                'width="100%" height="800px" style="border:1px solid var(--border);border-radius:8px;margin-top:12px;" ' +
                'type="application/pdf"></iframe>'
            )
            st.markdown(pdf_html, unsafe_allow_html=True)

    with tab_settings:
        display_settings_panel(data)


if __name__ == "__main__":
    main()
